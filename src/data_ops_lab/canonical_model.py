from __future__ import annotations

import csv
import hashlib
import re
import shutil
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import pandas as pd
import yaml

from .io_utils import normalize_columns, read_table
from .source_onboarding import backup_existing, ensure_dir


CONFIG_DIR = Path("config/data_model")
OUTPUT_DIR = Path("outputs/originaldatabase_analysis/step3e_human_approval_spreadsheet")
DATA_DIR = Path("originaldatabase")


@dataclass(frozen=True)
class CanonicalModelResult:
    config_dir: Path
    output_dir: Path
    canonical_count: int
    complement_count: int
    canonical_review_xlsx: Path
    product_status: str
    organisation_status: str


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def write_yaml(path: Path, payload: dict[str, Any], current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    path.write_text(yaml.safe_dump(payload, sort_keys=False, allow_unicode=True), encoding="utf-8")


def write_text(path: Path, text: str, current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    path.write_text(text, encoding="utf-8")


def canonical_tables_payload() -> dict[str, Any]:
    return {
        "generated_at": now_iso(),
        "canonical_tables": {
            "creditor": {
                "business_name": "Creditor / Supplier",
                "expected_prefix": "CR",
                "semantic_namespace": "creditor",
                "semantic_ref_name": "cr_ref_nr",
                "table_role": "canonical_master",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "customerproject": {
                "business_name": "Customer Project",
                "expected_prefix": "CP",
                "semantic_namespace": "customer_project",
                "semantic_ref_name": "cp_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "debtor": {
                "business_name": "Debtor / Customer",
                "expected_prefix": "DE",
                "semantic_namespace": "debtor",
                "semantic_ref_name": "de_ref_nr",
                "table_role": "canonical_master",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "deliverynote": {
                "business_name": "Delivery Note",
                "expected_prefix": "GU",
                "semantic_namespace": "delivery_note",
                "semantic_ref_name": "gu_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "goodsreception": {
                "business_name": "Goods Reception",
                "expected_prefix": "GO",
                "semantic_namespace": "goods_reception",
                "semantic_ref_name": "go_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "organisation": {
                "business_name": "Organisation",
                "expected_prefix": None,
                "semantic_namespace": "organisation",
                "semantic_ref_name": "organisation_ref_nr",
                "table_role": "canonical_master",
                "primary_key_rule": "generic_ref_nr_or_business_key",
                "status": "needs_business_context",
            },
            "product": {
                "business_name": "Product",
                "expected_prefix": "PD",
                "semantic_namespace": "product",
                "semantic_ref_name": "product_ref",
                "optional_serial_ref_name": "pd_ref_nr",
                "table_role": "canonical_master",
                "primary_key_rule": "part_nr_sku_business_reference",
                "optional_serial_rule": "PDYY99999 when available",
                "status": "manually_confirmed_pending_duplicate_validation",
            },
            "purchaseinvoice": {
                "business_name": "Purchase Invoice",
                "expected_prefix": "IF",
                "semantic_namespace": "purchase_invoice",
                "semantic_ref_name": "if_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "purchaseorder": {
                "business_name": "Purchase Order",
                "expected_prefix": "ON",
                "semantic_namespace": "purchase_order",
                "semantic_ref_name": "on_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "purchasequotation": {
                "business_name": "Purchase Quotation",
                "expected_prefix": "RFQ",
                "semantic_namespace": "purchase_quotation",
                "semantic_ref_name": "rfq_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "salesinvoice": {
                "business_name": "Sales Invoice",
                "expected_prefix": "CI",
                "semantic_namespace": "sales_invoice",
                "semantic_ref_name": "ci_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "salesopportunity": {
                "business_name": "Sales Opportunity",
                "expected_prefix": "VK",
                "semantic_namespace": "sales_opportunity",
                "semantic_ref_name": "vk_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "salesorder": {
                "business_name": "Sales Order",
                "expected_prefix": "OC",
                "semantic_namespace": "sales_order",
                "semantic_ref_name": "oc_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
            "salesquotation": {
                "business_name": "Sales Quotation",
                "expected_prefix": "CQ",
                "semantic_namespace": "sales_quotation",
                "semantic_ref_name": "cq_ref_nr",
                "table_role": "canonical_document",
                "primary_key_rule": "serial_ref_nr",
                "status": "manually_confirmed_pending_application",
            },
        },
    }


def business_rules_payload() -> dict[str, Any]:
    return {
        "generated_at": now_iso(),
        "canonical_table_rules": [
            "Canonical tables define the core data model only after manual approval is applied.",
            "Serial references remain pending until approved through the human review process.",
        ],
        "document_line_rules": [
            "A line table contains internal document lines for a header table.",
            "ref_nr in a line table points to the header document.",
            "ref_nr alone is not a primary key for line tables.",
            "ref_nr + row_position may be used only as a technical analytical key.",
        ],
        "product_rules": [
            "Product canonical reference is part_nr_sku.",
            "part_nr_sku is equivalent to Part. nr. (SKU).",
            "PD serial reference is optional, not mandatory.",
            "Products without PD code are valid product records.",
            "part_nr_sku should be validated as a business key candidate.",
            "pd_ref_nr should only be derived when the value matches PD serial format.",
            "If part_nr_sku starts with PD and matches PDYY99999, it can also be interpreted as pd_ref_nr.",
            "Product remains manually_confirmed_pending_duplicate_validation until duplicate validation is reviewed.",
        ],
        "organisation_rules": [
            "Organisation is a canonical master table.",
            "Organisation may use a generic ref_nr, not a serial prefix.",
            "Do not force a serial prefix for Organisation.",
            "Validate uniqueness and non-null rate for organisation.ref_nr or other business key candidates.",
            "Keep Organisation pending review until explicitly approved.",
        ],
        "complement_table_rules": [
            "Any table not listed in canonical_tables.yml starts as document_line, enrichment, complement, bridge_candidate, support_table, or needs_business_context.",
            "Complement tables should not define the core model unless manually approved.",
            "Not every field in complement tables is relevant to the core model.",
        ],
    }


def product_dataframe(data_dir: Path) -> pd.DataFrame:
    path = data_dir / "Product.xlsx"
    workbook = pd.ExcelFile(path)
    return normalize_columns(read_table(path, sheet_name=workbook.sheet_names[0]))


def read_matrix_rows(output_dir: Path) -> list[dict[str, str]]:
    path = output_dir / "human_approval_matrix.csv"
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def canonical_base(table_name: str) -> str:
    name = table_name.strip().lower()
    if not name:
        return ""
    if "_export_" in name:
        name = name.split("_export_", 1)[0]
    return re.sub(r"_?\d+$", "", name)


def is_canonical_physical_table(table_name: str, canonical_names: set[str]) -> bool:
    return canonical_base(table_name) in canonical_names


def complement_role(table_name: str) -> str:
    base = canonical_base(table_name)
    if "line" in base or "detail" in base:
        return "document_line"
    if "link" in base or "bridge" in base or "relation" in base:
        return "bridge_candidate"
    if "group" in base or "category" in base or "tag" in base:
        return "enrichment"
    if "setting" in base or "log" in base or "history" in base:
        return "support_table"
    return "complement"


def matrix_table_names(rows: list[dict[str, str]]) -> set[str]:
    names: set[str] = set()
    for row in rows:
        for key in ("table_name", "source_table", "target_table"):
            value = row.get(key, "").strip()
            if value:
                names.add(value)
    return names


def physical_tables_by_canonical(rows: list[dict[str, str]], canonical_names: set[str]) -> dict[str, list[str]]:
    mapping = {name: [] for name in canonical_names}
    for table_name in matrix_table_names(rows):
        base = canonical_base(table_name)
        if base in canonical_names:
            mapping[base].append(table_name)
    return {name: sorted(set(values)) for name, values in mapping.items()}


def complement_tables(rows: list[dict[str, str]], canonical_names: set[str]) -> list[dict[str, str]]:
    complements = []
    for table_name in sorted(matrix_table_names(rows)):
        if not is_canonical_physical_table(table_name, canonical_names):
            complements.append(
                {
                    "table_name": table_name,
                    "classification": complement_role(table_name),
                    "status": "pending_human_review",
                }
            )
    return complements


def masked_hash(value: str) -> str:
    return hashlib.sha256(value.encode("utf-8")).hexdigest()[:12]


def product_validation(df: pd.DataFrame) -> dict[str, Any]:
    series = df["part_nr_sku"].astype("string").str.strip()
    non_empty = series.dropna()
    non_empty = non_empty[non_empty != ""]
    pd_mask = non_empty.str.match(r"^PD[0-9]{2}[0-9]{5}$")
    duplicate_counts = non_empty.value_counts()
    duplicates = duplicate_counts[duplicate_counts > 1]
    duplicate_examples = []
    for idx, (value, count) in enumerate(duplicates.head(10).items(), start=1):
        duplicate_rows = df[series == value]
        exact_duplicate_rows = int(duplicate_rows.duplicated().sum())
        duplicate_examples.append(
            {
                "duplicate_group_id": f"product_dup_{idx:03d}_{masked_hash(str(value))}",
                "duplicate_count": int(count),
                "exact_duplicate_rows": exact_duplicate_rows,
                "different_product_records": bool(count - exact_duplicate_rows > 1),
            }
        )
    return {
        "total_products": int(len(df)),
        "part_nr_sku_filled": int(len(non_empty)),
        "part_nr_sku_empty": int(len(df) - len(non_empty)),
        "part_nr_sku_pd_pattern_count": int(pd_mask.sum()),
        "part_nr_sku_non_pd_pattern_count": int((~pd_mask).sum()),
        "duplicate_count": int(duplicates.sum() - len(duplicates)),
        "duplicate_group_count": int(len(duplicates)),
        "duplicate_examples": duplicate_examples,
    }


def render_product_report(stats: dict[str, Any]) -> str:
    lines = [
        "# Product Reference Validation",
        "",
        "Product business rule: `part_nr_sku` is the canonical functional product reference. `PDYY99999` remains optional and should only derive `pd_ref_nr` when the value matches the serial pattern.",
        "",
        f"- Total products: {stats['total_products']}",
        f"- `part_nr_sku` filled: {stats['part_nr_sku_filled']}",
        f"- `part_nr_sku` empty: {stats['part_nr_sku_empty']}",
        f"- `part_nr_sku` matching `PDYY99999`: {stats['part_nr_sku_pd_pattern_count']}",
        f"- `part_nr_sku` not matching PD pattern: {stats['part_nr_sku_non_pd_pattern_count']}",
        f"- Duplicate count for `part_nr_sku`: {stats['duplicate_count']}",
        f"- Duplicate group count for `part_nr_sku`: {stats['duplicate_group_count']}",
        "",
        "## Duplicate Group Examples",
        "",
        "Values are not exposed directly; duplicate groups are represented by stable hash IDs.",
    ]
    if not stats["duplicate_examples"]:
        lines.append("- No duplicate groups detected.")
    else:
        for example in stats["duplicate_examples"]:
            lines.append(
                "- {duplicate_group_id}: duplicate_count={duplicate_count}, exact_duplicate_rows={exact_duplicate_rows}, different_product_records={different_product_records}".format(
                    **example
                )
            )
    lines.extend(
        [
            "",
            "## Recommendation For Human Approval",
            "",
            "- Keep `product_export_product.part_nr_sku` as `manually_confirmed_pending_duplicate_validation`.",
            "- Do not approve it as final primary key until duplicate groups are reviewed.",
            "- Treat `pd_ref_nr` as optional and derived only for values matching `PDYY99999`.",
        ]
    )
    return "\n".join(lines)


def render_alignment_report(
    payload: dict[str, Any],
    stats: dict[str, Any],
    rows: list[dict[str, str]],
    complements: list[dict[str, str]],
) -> str:
    canonical_names = set(payload["canonical_tables"])
    physical_mapping = physical_tables_by_canonical(rows, canonical_names)
    lines = [
        "# Canonical Model Alignment Report",
        "",
        "No approvals are applied by this report.",
        "",
        "## Canonical Tables",
    ]
    for name, table in payload["canonical_tables"].items():
        physical_tables = ", ".join(physical_mapping.get(name, [])) or "not identified in approval matrix"
        lines.append(
            f"- {name}: prefix={table.get('expected_prefix')}, physical_tables={physical_tables}, namespace={table['semantic_namespace']}, role={table['table_role']}, status={table['status']}"
        )
    lines.extend(["", "## Complement Tables", ""])
    if not complements:
        lines.append("- No complement tables detected from the current approval matrix.")
    else:
        for complement in complements:
            lines.append(
                f"- {complement['table_name']}: classification={complement['classification']}, status={complement['status']}"
            )
    lines.extend(
        [
            "",
            "## Product Status",
            "",
            "- Product semantic namespace: product",
            "- Product canonical reference: part_nr_sku",
            "- Product semantic ref name: product_ref",
            "- Optional serial ref name: pd_ref_nr",
            f"- Duplicate count pending review: {stats['duplicate_count']}",
            "- Status: manually_confirmed_pending_duplicate_validation",
            "",
            "## Pending Questions",
            "",
            "- Which duplicate product groups are valid repeated records versus data-quality issues?",
            "- Should duplicate product references be resolved before applying Product as a primary key?",
            "- Should non-PD product references remain as product_ref only?",
        ]
    )
    return "\n".join(lines)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)


def replace_sheet(workbook, sheet_name: str):
    if sheet_name in workbook.sheetnames:
        del workbook[sheet_name]
    return workbook.create_sheet(sheet_name)


def write_canonical_review_workbook(
    output_dir: Path,
    payload: dict[str, Any],
    rows: list[dict[str, str]],
    complements: list[dict[str, str]],
    stats: dict[str, Any],
    current_run_id: str,
) -> Path:
    ensure_dir(output_dir)
    source = output_dir / "human_approval_matrix.xlsx"
    target = output_dir / "human_approval_matrix_canonical_review.xlsx"
    backup_existing(target, current_run_id)
    if source.exists():
        shutil.copy2(source, target)
        workbook = load_workbook(target)
    else:
        workbook = Workbook()
        workbook.active.title = "README"
        workbook["README"].append(["Canonical review workbook. No approvals are applied."])

    canonical_names = set(payload["canonical_tables"])
    physical_mapping = physical_tables_by_canonical(rows, canonical_names)
    canonical_sheet = replace_sheet(workbook, "Canonical Model")
    canonical_sheet.append(
        [
            "canonical_table",
            "business_name",
            "expected_prefix",
            "physical_tables_identified",
            "semantic_namespace",
            "semantic_ref_name",
            "optional_serial_ref_name",
            "table_role",
            "primary_key_rule",
            "status",
        ]
    )
    for name, table in payload["canonical_tables"].items():
        canonical_sheet.append(
            [
                name,
                table["business_name"],
                table.get("expected_prefix"),
                ", ".join(physical_mapping.get(name, [])),
                table["semantic_namespace"],
                table["semantic_ref_name"],
                table.get("optional_serial_ref_name", ""),
                table["table_role"],
                table["primary_key_rule"],
                table["status"],
            ]
        )
    style_sheet(canonical_sheet)

    complement_sheet = replace_sheet(workbook, "Complement Tables")
    complement_sheet.append(["table_name", "classification", "status"])
    for complement in complements:
        complement_sheet.append([complement["table_name"], complement["classification"], complement["status"]])
    style_sheet(complement_sheet)

    product_sheet = replace_sheet(workbook, "Product Validation")
    product_sheet.append(["metric", "value"])
    for key in [
        "total_products",
        "part_nr_sku_filled",
        "part_nr_sku_empty",
        "part_nr_sku_pd_pattern_count",
        "part_nr_sku_non_pd_pattern_count",
        "duplicate_count",
        "duplicate_group_count",
    ]:
        product_sheet.append([key, stats[key]])
    product_sheet.append(["status", "manually_confirmed_pending_duplicate_validation"])
    product_sheet.append(["approval_note", "Do not approve part_nr_sku as final primary key until duplicate groups are reviewed."])
    style_sheet(product_sheet)

    workbook.save(target)
    return target


def run_canonical_model_alignment(
    data_dir: Path = DATA_DIR,
    config_dir: Path = CONFIG_DIR,
    output_dir: Path = OUTPUT_DIR,
) -> CanonicalModelResult:
    current_run_id = run_id()
    canonical_payload = canonical_tables_payload()
    rules_payload = business_rules_payload()
    df = product_dataframe(data_dir)
    stats = product_validation(df)
    matrix_rows = read_matrix_rows(output_dir)
    complements = complement_tables(matrix_rows, set(canonical_payload["canonical_tables"]))

    write_yaml(config_dir / "canonical_tables.yml", canonical_payload, current_run_id)
    write_yaml(config_dir / "business_rules.yml", rules_payload, current_run_id)
    write_text(output_dir / "product_reference_validation.md", render_product_report(stats), current_run_id)
    write_text(
        output_dir / "canonical_model_alignment_report.md",
        render_alignment_report(canonical_payload, stats, matrix_rows, complements),
        current_run_id,
    )
    canonical_review_xlsx = write_canonical_review_workbook(
        output_dir, canonical_payload, matrix_rows, complements, stats, current_run_id
    )

    return CanonicalModelResult(
        config_dir=config_dir,
        output_dir=output_dir,
        canonical_count=len(canonical_payload["canonical_tables"]),
        complement_count=len(complements),
        canonical_review_xlsx=canonical_review_xlsx,
        product_status=canonical_payload["canonical_tables"]["product"]["status"],
        organisation_status=canonical_payload["canonical_tables"]["organisation"]["status"],
    )
