from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .product_reference_audit import OUTPUT_DIR, PD_PATTERN, clean_value
from .source_onboarding import backup_existing, ensure_dir


RECONCILIATION_WORKBOOK = "product_refnr_reconciliation_review.xlsx"
SHORTLIST_WORKBOOK = "product_refnr_human_review_shortlist.xlsx"
SHORTLIST_REPORT = "product_refnr_human_review_shortlist.md"
MODELING_RECOMMENDATION = "product_modeling_recommendation.md"

DECISION_OPTIONS = [
    "approved_use_corrected_product_ref_nr",
    "approved_keep_original_part_nr_sku_only",
    "approved_create_technical_product_id_only",
    "merge_duplicate_records",
    "keep_as_separate_products",
    "needs_business_context",
    "rejected",
    "pending",
]

REVIEW_COLUMNS = [
    "issue_id",
    "issue_type",
    "product_original_identifier",
    "product_refnr_identifier",
    "original_part_nr_sku",
    "corrected_product_ref_nr",
    "optional_pd_ref_nr",
    "match_reason",
    "conflict_reason",
    "risk_explanation",
    "recommended_action",
    "human_decision",
    "human_notes",
]


@dataclass(frozen=True)
class ProductRefnrHumanReviewResult:
    output_dir: Path
    shortlist_xlsx: Path
    shortlist_md: Path
    modeling_recommendation_path: Path
    conflict_count: int
    unmatched_original_count: int
    unmatched_refnr_count: int
    duplicate_count: int
    product_finalized: bool


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def rows_from_sheet(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_value(value) for value in rows[0]]
    output = []
    for raw_row in rows[1:]:
        if not any(clean_value(value) for value in raw_row):
            continue
        output.append({header: clean_value(value) for header, value in zip(headers, raw_row, strict=False) if header})
    return output


def optional_pd(value: str) -> str:
    return value if PD_PATTERN.match(value) else ""


def original_identifier(row: dict[str, Any]) -> str:
    source_row = clean_value(row.get("original_source_row_number", ""))
    if source_row:
        return f"original_row_{source_row}"
    return ""


def refnr_identifier(row: dict[str, Any]) -> str:
    source_row = clean_value(row.get("refnr_source_row_number", ""))
    if source_row:
        return f"refnr_row_{source_row}"
    return ""


def issue_row(
    issue_id: str,
    issue_type: str,
    row: dict[str, Any],
    recommended_action: str,
    risk_explanation: str,
    conflict_reason: str = "",
    match_reason: str = "",
) -> dict[str, Any]:
    corrected_ref = clean_value(row.get("refnr_product_ref_nr", "")) or clean_value(row.get("corrected_ref_nr", ""))
    original_part = clean_value(row.get("original_part_nr_sku", "")) or clean_value(row.get("refnr_part_nr_sku", ""))
    return {
        "issue_id": issue_id,
        "issue_type": issue_type,
        "product_original_identifier": original_identifier(row),
        "product_refnr_identifier": refnr_identifier(row),
        "original_part_nr_sku": original_part,
        "corrected_product_ref_nr": corrected_ref,
        "optional_pd_ref_nr": optional_pd(corrected_ref),
        "match_reason": match_reason or clean_value(row.get("match_method", "")),
        "conflict_reason": conflict_reason or clean_value(row.get("conflict_reason", "")),
        "risk_explanation": risk_explanation,
        "recommended_action": recommended_action,
        "human_decision": "pending",
        "human_notes": "",
    }


def build_conflict_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        issue_row(
            f"CONFLICT_{idx:03d}",
            "conflict",
            row,
            "needs_business_context",
            "Multiple or missing reconciliation candidates can create incorrect Product reference assignment.",
        )
        for idx, row in enumerate(rows, start=1)
    ]


def build_unmatched_original_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        issue_row(
            f"UNMATCHED_ORIGINAL_{idx:03d}",
            "unmatched_original_product",
            row,
            "approved_create_technical_product_id_only",
            "Original Product row has no matched corrected Product_ref.nr row and must be reviewed before canonical reference assignment.",
            conflict_reason="no_matching_product_refnr_row",
        )
        for idx, row in enumerate(rows, start=1)
    ]


def build_unmatched_refnr_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        issue_row(
            f"UNMATCHED_REFNR_{idx:03d}",
            "unmatched_product_refnr",
            row,
            "needs_business_context",
            "Product_ref.nr row did not match the original Product export and may represent a new, duplicate, obsolete, or unmatched product.",
            conflict_reason="no_matching_original_product_row",
        )
        for idx, row in enumerate(rows, start=1)
    ]


def build_duplicate_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    output = []
    for idx, row in enumerate(rows, start=1):
        duplicate_column = clean_value(row.get("duplicate_column", ""))
        duplicate_count = clean_value(row.get("duplicate_count", ""))
        output.append(
            issue_row(
                f"DUPLICATE_REFNR_{idx:03d}",
                "duplicate_refnr_review",
                row,
                "needs_business_context",
                "Duplicate values in Product_ref.nr need human confirmation before canonical reference rules are finalized.",
                conflict_reason=f"duplicate {duplicate_column}; duplicate_count={duplicate_count}",
            )
        )
    return output


def style_sheet(sheet, add_validation: bool = False) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    human_decision_col = None
    for cell in sheet[1]:
        if cell.value in {"human_decision", "human_notes"}:
            letter = get_column_letter(cell.column)
            for data_cell in sheet[letter][1:]:
                data_cell.fill = input_fill
        if cell.value == "human_decision":
            human_decision_col = cell.column
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 14), 55)
    if add_validation and human_decision_col:
        formula = '"' + ",".join(DECISION_OPTIONS) + '"'
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = "Choose one of the allowed Product reconciliation decisions."
        validation.errorTitle = "Invalid decision"
        sheet.add_data_validation(validation)
        letter = get_column_letter(human_decision_col)
        validation.add(f"{letter}2:{letter}{max(sheet.max_row, 2)}")


def write_review_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(REVIEW_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in REVIEW_COLUMNS])
    style_sheet(sheet, add_validation=True)


def write_readme(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("README")
    sheet.append(["Product Reconciliation Human Review"])
    sheet.append(["This workbook contains only reconciliation exceptions. No Product key approval has been applied."])
    sheet.append(["Review each issue sheet, choose human_decision, and add human_notes where needed."])
    sheet.append(["Final Product key approval remains pending until this shortlist is reviewed."])
    sheet["A1"].font = Font(bold=True, size=16)
    sheet.column_dimensions["A"].width = 120


def write_decision_options(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Decision Options")
    sheet.append(["human_decision", "description"])
    descriptions = {
        "approved_use_corrected_product_ref_nr": "Use corrected product_ref_nr from Product_ref.nr for this Product row.",
        "approved_keep_original_part_nr_sku_only": "Keep only the original part_nr_sku business reference for this exception.",
        "approved_create_technical_product_id_only": "Use generated product_id only and leave canonical reference unresolved for now.",
        "merge_duplicate_records": "Treat duplicate rows as the same product and merge/clean them before import.",
        "keep_as_separate_products": "Keep rows as separate products even if references overlap.",
        "needs_business_context": "Business owner must clarify the correct treatment.",
        "rejected": "Reject the proposed reconciliation for this row.",
        "pending": "No final human decision has been made.",
    }
    for option in DECISION_OPTIONS:
        sheet.append([option, descriptions[option]])
    style_sheet(sheet)


def write_workbook(path: Path, conflicts: list[dict[str, Any]], unmatched_original: list[dict[str, Any]], unmatched_refnr: list[dict[str, Any]], duplicates: list[dict[str, Any]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_readme(workbook)
    write_review_sheet(workbook, "Conflicts", conflicts)
    write_review_sheet(workbook, "Unmatched Original Product", unmatched_original)
    write_review_sheet(workbook, "Unmatched Product RefNr", unmatched_refnr)
    write_review_sheet(workbook, "Duplicate RefNr Review", duplicates)
    write_decision_options(workbook)
    workbook.save(path)


def render_shortlist_report(conflicts: list[dict[str, Any]], unmatched_original: list[dict[str, Any]], unmatched_refnr: list[dict[str, Any]], duplicates: list[dict[str, Any]]) -> str:
    lines = [
        "# Product RefNr Human Review Shortlist",
        "",
        "This shortlist contains only Product reconciliation exceptions. No approval is applied automatically.",
        "",
        "## Summary",
        "",
        f"- Conflicts: {len(conflicts)}",
        f"- Unmatched original Product rows: {len(unmatched_original)}",
        f"- Unmatched Product_ref.nr rows: {len(unmatched_refnr)}",
        f"- Duplicate Product_ref.nr review rows: {len(duplicates)}",
        "",
        "## Questions For Human Decision",
        "",
        "- For conflicts, which corrected `product_ref_nr` should be used, if any?",
        "- For unmatched original Product rows, should they be excluded, assigned only a technical `product_id`, or manually mapped?",
        "- For unmatched Product_ref.nr rows, are they new valid products, obsolete references, or unmatched duplicates?",
        "- For duplicate Product_ref.nr rows, should records be merged or kept as separate products?",
        "",
        "## Modeling Recommendation Pending",
        "",
        "- Product remains a canonical master table.",
        "- Final Product key approval remains pending until this reconciliation review is completed.",
        "- Proposed final model remains `product_id` as technical primary key, `product_ref_nr` as corrected canonical ERP reference, `part_nr_sku` as business/search reference, and optional `pd_ref_nr` when available.",
    ]
    return "\n".join(lines)


def render_modeling_recommendation() -> str:
    return "\n".join(
        [
            "# Product Modeling Recommendation",
            "",
            "- Product should use `product_id` as technical/generated primary key.",
            "- `product_ref_nr` should be the canonical corrected ERP product reference when reconciled.",
            "- `part_nr_sku` is business/search/customer/supplier-facing reference.",
            "- `pd_ref_nr` is optional and only populated when a PD-style reference exists.",
            "- Final approval remains pending until the reconciliation review is completed.",
            "- Do not update `approved_keys.yml` or `approved_relationships.yml` from this recommendation.",
        ]
    )


def run_product_refnr_human_review(
    output_dir: Path = OUTPUT_DIR,
    reconciliation_workbook: Path | None = None,
) -> ProductRefnrHumanReviewResult:
    current_run_id = run_id()
    reconciliation_workbook = reconciliation_workbook or output_dir / RECONCILIATION_WORKBOOK
    if not reconciliation_workbook.exists():
        raise FileNotFoundError(f"Product reconciliation workbook not found: {reconciliation_workbook}")
    workbook = load_workbook(reconciliation_workbook, data_only=True)

    conflicts = build_conflict_rows(rows_from_sheet(workbook, "conflicts"))
    unmatched_original = build_unmatched_original_rows(rows_from_sheet(workbook, "unmatched_original_product"))
    unmatched_refnr = build_unmatched_refnr_rows(rows_from_sheet(workbook, "unmatched_product_refnr"))
    duplicates = build_duplicate_rows(rows_from_sheet(workbook, "duplicates_in_product_refnr"))

    ensure_dir(output_dir)
    shortlist_xlsx = output_dir / SHORTLIST_WORKBOOK
    shortlist_md = output_dir / SHORTLIST_REPORT
    modeling_recommendation_path = output_dir / MODELING_RECOMMENDATION
    backup_existing(shortlist_xlsx, current_run_id)
    backup_existing(shortlist_md, current_run_id)
    backup_existing(modeling_recommendation_path, current_run_id)
    write_workbook(shortlist_xlsx, conflicts, unmatched_original, unmatched_refnr, duplicates)
    shortlist_md.write_text(render_shortlist_report(conflicts, unmatched_original, unmatched_refnr, duplicates), encoding="utf-8")
    modeling_recommendation_path.write_text(render_modeling_recommendation(), encoding="utf-8")

    return ProductRefnrHumanReviewResult(
        output_dir=output_dir,
        shortlist_xlsx=shortlist_xlsx,
        shortlist_md=shortlist_md,
        modeling_recommendation_path=modeling_recommendation_path,
        conflict_count=len(conflicts),
        unmatched_original_count=len(unmatched_original),
        unmatched_refnr_count=len(unmatched_refnr),
        duplicate_count=len(duplicates),
        product_finalized=False,
    )
