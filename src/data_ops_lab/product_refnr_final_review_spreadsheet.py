from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .product_reference_audit import OUTPUT_DIR, clean_value
from .product_refnr_decision_validation import ALLOWED_DECISIONS, REVIEW_SHEETS, inconsistency, normalized_decision, note, requires_note
from .product_refnr_human_review import DECISION_OPTIONS, SHORTLIST_WORKBOOK
from .source_onboarding import backup_existing, ensure_dir


XLSX_NAME = "product_refnr_final_review_required.xlsx"
CSV_NAME = "product_refnr_final_review_required.csv"
README_NAME = "product_refnr_final_review_required_readme.md"

REVIEW_COLUMNS = [
    "review_id",
    "original_sheet",
    "original_excel_row",
    "issue_id",
    "issue_type",
    "product_original_identifier",
    "product_refnr_identifier",
    "original_part_nr_sku",
    "corrected_product_ref_nr",
    "optional_pd_ref_nr",
    "current_human_decision",
    "current_human_notes",
    "problem_type",
    "problem_explanation",
    "suggested_human_decision",
    "suggested_human_notes",
    "required_action",
    "final_human_decision",
    "final_human_notes",
]

NOTE_SUGGESTIONS = {
    "approved_use_corrected_product_ref_nr": "Corrected product reference accepted after manual review.",
    "approved_keep_original_part_nr_sku_only": "No corrected Product_ref.nr match found; original part_nr_sku kept as functional product reference.",
    "approved_create_technical_product_id_only": "No reliable natural product reference confirmed; product retained with generated technical product_id only.",
    "merge_duplicate_records": "Duplicate records confirmed as the same product after manual review.",
    "keep_as_separate_products": "Records confirmed as separate products after manual review.",
    "needs_business_context": "Requires additional business context before final product modeling decision.",
}


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def rows_from_sheet_with_index(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_value(value) for value in rows[0]]
    output = []
    for excel_row, raw_row in enumerate(rows[1:], start=2):
        if not any(clean_value(value) for value in raw_row):
            continue
        row = {header: clean_value(value) for header, value in zip(headers, raw_row, strict=False) if header}
        row["_original_sheet"] = sheet_name
        row["_original_excel_row"] = excel_row
        output.append(row)
    return output


def issue_key(row: dict[str, Any]) -> tuple[str, int, str]:
    return (
        clean_value(row.get("_original_sheet", "")),
        int(row.get("_original_excel_row", 0)),
        clean_value(row.get("issue_id", "")),
    )


def suggested_note_for_decision(decision: str) -> str:
    return NOTE_SUGGESTIONS.get(decision, "")


def inconsistency_explanation(problem: str) -> str:
    explanations = {
        "conflict_keeps_original_without_corrected_mapping": "The issue is a reconciliation conflict, but the current decision keeps the original part_nr_sku without resolving the corrected Product_ref.nr mapping.",
        "unmatched_original_has_no_corrected_refnr_to_use": "The original Product row has no matched Product_ref.nr row, so the current decision cannot use a corrected Product_ref.nr.",
        "unmatched_refnr_has_no_original_product_to_keep": "The Product_ref.nr row has no matching original Product row, so keeping the original part_nr_sku only is not applicable.",
        "duplicate_refnr_approved_without_duplicate_resolution_note": "A duplicate Product_ref.nr row was approved without a note explaining whether records should be merged or kept separate.",
    }
    return explanations.get(problem, "The current human decision does not align cleanly with the issue type or recommended action.")


def inconsistency_action(problem: str) -> str:
    if problem == "duplicate_refnr_approved_without_duplicate_resolution_note":
        return "Add human_notes explaining duplicate resolution, or change human_decision to merge_duplicate_records / keep_as_separate_products."
    return "Review human_decision and human_notes; adjust one or both before applying final Product reconciliation decisions."


def problem_types_for_row(row: dict[str, Any]) -> list[str]:
    problems = []
    decision = normalized_decision(row)
    if not decision or decision not in ALLOWED_DECISIONS:
        problems.append("invalid_or_empty_decision")
    elif decision == "pending":
        problems.append("pending")
    if requires_note(row) and not note(row):
        problems.append("missing_note")
    if inconsistency(row):
        problems.append("inconsistency")
    return problems


def problem_explanation(row: dict[str, Any], problems: list[str]) -> str:
    parts = []
    decision = normalized_decision(row)
    if "missing_note" in problems:
        parts.append(f"`{decision}` requires a human note for this issue type.")
    if "inconsistency" in problems:
        parts.append(inconsistency_explanation(inconsistency(row)))
    if "invalid_or_empty_decision" in problems:
        parts.append("The human_decision value is empty or not one of the allowed options.")
    if "pending" in problems:
        parts.append("The human_decision is still pending.")
    return " ".join(parts) or "No blocking issue detected."


def required_action(row: dict[str, Any], problems: list[str]) -> str:
    actions = []
    if "missing_note" in problems:
        actions.append("Fill final_human_notes.")
    if "inconsistency" in problems:
        actions.append(inconsistency_action(inconsistency(row)))
    if "invalid_or_empty_decision" in problems or "pending" in problems:
        actions.append("Choose a valid final_human_decision.")
    return " ".join(actions) or "No action required."


def review_row(row: dict[str, Any], review_id: str, problem_type: str, problems: list[str]) -> dict[str, Any]:
    decision = normalized_decision(row)
    current_notes = note(row)
    return {
        "review_id": review_id,
        "original_sheet": row["_original_sheet"],
        "original_excel_row": row["_original_excel_row"],
        "issue_id": clean_value(row.get("issue_id", "")),
        "issue_type": clean_value(row.get("issue_type", "")),
        "product_original_identifier": clean_value(row.get("product_original_identifier", "")),
        "product_refnr_identifier": clean_value(row.get("product_refnr_identifier", "")),
        "original_part_nr_sku": clean_value(row.get("original_part_nr_sku", "")),
        "corrected_product_ref_nr": clean_value(row.get("corrected_product_ref_nr", "")),
        "optional_pd_ref_nr": clean_value(row.get("optional_pd_ref_nr", "")),
        "current_human_decision": decision,
        "current_human_notes": current_notes,
        "problem_type": problem_type,
        "problem_explanation": problem_explanation(row, problems),
        "suggested_human_decision": decision if decision in ALLOWED_DECISIONS else "pending",
        "suggested_human_notes": suggested_note_for_decision(decision),
        "required_action": required_action(row, problems),
        "final_human_decision": decision if decision in ALLOWED_DECISIONS else "pending",
        "final_human_notes": current_notes,
    }


def collect_review_data(workbook) -> dict[str, list[dict[str, Any]]]:
    all_rows = []
    for sheet in REVIEW_SHEETS:
        all_rows.extend(rows_from_sheet_with_index(workbook, sheet))

    required = []
    missing = []
    inconsistencies = []
    all_exceptions = []
    for index, row in enumerate(all_rows, start=1):
        problems = problem_types_for_row(row)
        problem_type = ";".join(problems) if problems else "ready"
        rendered = review_row(row, f"REVIEW_{index:03d}", problem_type, problems)
        all_exceptions.append(rendered)
        if problems:
            required.append(rendered)
        if "missing_note" in problems:
            missing.append(rendered)
        if "inconsistency" in problems:
            inconsistencies.append(rendered)
    return {
        "required": required,
        "missing": missing,
        "inconsistencies": inconsistencies,
        "all": all_exceptions,
    }


def write_csv(path: Path, rows: list[dict[str, Any]], current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=REVIEW_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def style_sheet(sheet, add_validation: bool = False) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    fills = {
        "missing_note": PatternFill("solid", fgColor="FCE4D6"),
        "inconsistency": PatternFill("solid", fgColor="F4CCCC"),
        "ready": PatternFill("solid", fgColor="E2F0D9"),
        "pending": PatternFill("solid", fgColor="FFF2CC"),
        "invalid_or_empty_decision": PatternFill("solid", fgColor="EADCF8"),
    }
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions

    headers = [cell.value for cell in sheet[1]]
    problem_col = headers.index("problem_type") + 1 if "problem_type" in headers else None
    decision_col = headers.index("final_human_decision") + 1 if "final_human_decision" in headers else None
    notes_col = headers.index("final_human_notes") + 1 if "final_human_notes" in headers else None
    for row in sheet.iter_rows(min_row=2):
        row_fill = None
        if problem_col:
            problem_value = str(sheet.cell(row=row[0].row, column=problem_col).value or "")
            for key, fill in fills.items():
                if key in problem_value:
                    row_fill = fill
                    break
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_fill:
                cell.fill = row_fill
        for column_index in (decision_col, notes_col):
            if column_index:
                sheet.cell(row=row[0].row, column=column_index).fill = input_fill

    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 14), 60)

    if add_validation and decision_col:
        formula = "'Decision Options'!$A$2:$A$9"
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        validation.error = "Choose one of the allowed Product reconciliation decisions."
        validation.errorTitle = "Invalid decision"
        sheet.add_data_validation(validation)
        letter = get_column_letter(decision_col)
        validation.add(f"{letter}2:{letter}{max(sheet.max_row, 2)}")


def write_review_sheet(workbook: Workbook, title: str, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(REVIEW_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in REVIEW_COLUMNS])
    style_sheet(sheet, add_validation=True)


def write_readme_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("README")
    rows = [
        ["Product Final Review Spreadsheet"],
        ["This workbook contains only the items that still block the final Product decision."],
        ["No decision will be applied automatically from this workbook."],
        ["Fill or correct final_human_decision and final_human_notes in the review sheets."],
        ["Run validation again after completing the corrections."],
        ["Only after a clean validation should Step 3E.4 apply Product reconciliation decisions."],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(bold=True, size=16)
    sheet.column_dimensions["A"].width = 120


def write_decision_options(workbook: Workbook) -> None:
    descriptions = {
        "approved_use_corrected_product_ref_nr": "Use corrected product_ref_nr from Product_ref.nr.",
        "approved_keep_original_part_nr_sku_only": "Keep original part_nr_sku only as the functional business reference.",
        "approved_create_technical_product_id_only": "Use generated product_id only; no reliable natural reference confirmed.",
        "merge_duplicate_records": "Merge records confirmed as the same Product.",
        "keep_as_separate_products": "Keep records as separate Products.",
        "needs_business_context": "Requires additional business clarification.",
        "rejected": "Reject the proposed reconciliation decision.",
        "pending": "No final decision has been made.",
    }
    sheet = workbook.create_sheet("Decision Options")
    sheet.append(["human_decision", "description"])
    for option in DECISION_OPTIONS:
        sheet.append([option, descriptions[option]])
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.font = Font(bold=True, color="FFFFFF")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    sheet.column_dimensions["A"].width = 42
    sheet.column_dimensions["B"].width = 95


def write_workbook(path: Path, data: dict[str, list[dict[str, Any]]], current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_readme_sheet(workbook)
    write_review_sheet(workbook, "Required Review", data["required"])
    write_review_sheet(workbook, "Missing Notes", data["missing"])
    write_review_sheet(workbook, "Inconsistencies", data["inconsistencies"])
    write_review_sheet(workbook, "All Product Exceptions", data["all"])
    write_decision_options(workbook)
    workbook.save(path)


def render_readme(data: dict[str, list[dict[str, Any]]]) -> str:
    return "\n".join(
        [
            "# Product RefNr Final Review Required",
            "",
            "This workbook contains only Product reconciliation items that still block final application.",
            "",
            f"- Required Review rows: {len(data['required'])}",
            f"- Missing Notes rows: {len(data['missing'])}",
            f"- Inconsistency rows: {len(data['inconsistencies'])}",
            f"- All Product Exceptions rows: {len(data['all'])}",
            "",
            "No decisions are applied automatically. Complete `final_human_decision` and `final_human_notes`, then run `validate-product-refnr-decisions` again before Step 3E.4.",
        ]
    )


@dataclass(frozen=True)
class ProductRefnrFinalReviewSpreadsheetResult:
    output_dir: Path
    xlsx_path: Path
    csv_path: Path
    readme_path: Path
    required_review_count: int
    missing_notes_count: int
    inconsistency_count: int
    all_exceptions_count: int


def run_product_refnr_final_review_spreadsheet(
    output_dir: Path = OUTPUT_DIR,
    shortlist_path: Path | None = None,
) -> ProductRefnrFinalReviewSpreadsheetResult:
    current_run_id = run_id()
    shortlist_path = shortlist_path or output_dir / SHORTLIST_WORKBOOK
    if not shortlist_path.exists():
        raise FileNotFoundError(f"Product RefNr shortlist not found: {shortlist_path}")
    workbook = load_workbook(shortlist_path, data_only=True)
    data = collect_review_data(workbook)

    xlsx_path = output_dir / XLSX_NAME
    csv_path = output_dir / CSV_NAME
    readme_path = output_dir / README_NAME
    write_workbook(xlsx_path, data, current_run_id)
    write_csv(csv_path, data["required"], current_run_id)
    backup_existing(readme_path, current_run_id)
    readme_path.write_text(render_readme(data), encoding="utf-8")

    return ProductRefnrFinalReviewSpreadsheetResult(
        output_dir=output_dir,
        xlsx_path=xlsx_path,
        csv_path=csv_path,
        readme_path=readme_path,
        required_review_count=len(data["required"]),
        missing_notes_count=len(data["missing"]),
        inconsistency_count=len(data["inconsistencies"]),
        all_exceptions_count=len(data["all"]),
    )
