from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .product_reference_audit import (
    DATA_DIR,
    OUTPUT_DIR,
    PD_PATTERN,
    audit_stats,
    clean_value,
    duplicate_group_audit,
    empty_reference_audit,
    product_dataframe,
    reference_series,
)
from .source_onboarding import backup_existing, ensure_dir


INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")

DUPLICATE_DECISIONS = [
    "same_product_duplicate_record",
    "distinct_products_same_reference",
    "incorrect_reference",
    "obsolete_or_inactive_record",
    "requires_more_investigation",
]
EMPTY_DECISIONS = [
    "exclude_from_product_master",
    "repair_reference",
    "keep_as_exception",
    "requires_more_investigation",
]
NON_PD_DECISIONS = [
    "valid_textual_product_reference",
    "should_have_pd_reference",
    "incorrect_reference",
    "obsolete_or_inactive_record",
    "requires_more_investigation",
]


@dataclass(frozen=True)
class ProductReferenceReviewSpreadsheetResult:
    output_dir: Path
    xlsx_path: Path
    duplicate_rows: int
    empty_rows: int
    non_pd_rows: int
    rows_requiring_human_review: int


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def source_row_number(index: Any) -> int:
    return int(index) + 2


def pd_ref_nr(value: str) -> str:
    return value if PD_PATTERN.match(value) else ""


def relevant_product_columns(df: pd.DataFrame) -> list[str]:
    hints = (
        "description",
        "name",
        "supplier",
        "suppl",
        "creditor",
        "group",
        "category",
        "status",
        "active",
        "price",
        "cost",
        "sales",
        "date",
        "barcode",
        "hour_type",
    )
    columns = []
    for column in df.columns:
        if column == "part_nr_sku":
            continue
        if any(hint in column.lower() for hint in hints):
            columns.append(column)
    return columns


def row_payload(row: pd.Series, columns: list[str]) -> dict[str, Any]:
    return {column: clean_value(row.get(column, "")) for column in columns}


def duplicate_rows(df: pd.DataFrame, groups: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    refs = reference_series(df)
    group_by_raw_ref = {}
    duplicate_counts = refs[refs != ""].value_counts()
    duplicate_values = duplicate_counts[duplicate_counts > 1]
    for group, raw_value in zip(groups, duplicate_values.index, strict=False):
        group_by_raw_ref[raw_value] = group

    rows: list[dict[str, Any]] = []
    for raw_value, group in group_by_raw_ref.items():
        for index, row in df[refs == raw_value].iterrows():
            part_ref = clean_value(row["part_nr_sku"])
            rows.append(
                {
                    "duplicate_group_id": group["group_id"],
                    "source_row_number": source_row_number(index),
                    "part_nr_sku": part_ref,
                    "pd_ref_nr": pd_ref_nr(part_ref),
                    **row_payload(row, columns),
                    "classification_from_audit": group["classification"],
                    "suggested_review_question": "Are these rows the same product, distinct products sharing one reference, or an incorrect reference?",
                    "human_decision": "",
                    "human_notes": "",
                }
            )
    return rows


def empty_rows(df: pd.DataFrame, audits: list[dict[str, Any]], columns: list[str]) -> list[dict[str, Any]]:
    refs = reference_series(df)
    empty_df = df[refs == ""]
    rows = []
    for audit, (index, row) in zip(audits, empty_df.iterrows(), strict=False):
        rows.append(
            {
                "source_row_number": source_row_number(index),
                "pd_ref_nr": "",
                **row_payload(row, columns),
                "current_audit_classification": audit["classification"],
                "suggested_action": suggested_empty_action(audit["classification"]),
                "human_decision": "",
                "human_notes": "",
            }
        )
    return rows


def non_pd_rows(df: pd.DataFrame, columns: list[str]) -> list[dict[str, Any]]:
    refs = reference_series(df)
    mask = (refs != "") & ~refs.map(lambda value: bool(PD_PATTERN.match(value)))
    rows = []
    for index, row in df[mask].iterrows():
        part_ref = clean_value(row["part_nr_sku"])
        rows.append(
            {
                "source_row_number": source_row_number(index),
                "part_nr_sku": part_ref,
                "pd_ref_nr": "",
                **row_payload(row, columns),
                "suggested_interpretation": "valid_textual_product_reference_pending_human_review",
                "human_decision": "",
                "human_notes": "",
            }
        )
    return rows


def suggested_empty_action(classification: str) -> str:
    if classification == "can_be_excluded":
        return "exclude_from_product_master"
    if classification == "repair_candidate":
        return "repair_reference"
    return "requires_more_investigation"


def rows_requiring_human_review(duplicates: list[dict[str, Any]], empties: list[dict[str, Any]], non_pd: list[dict[str, Any]]) -> int:
    empty_review = sum(1 for row in empties if row["current_audit_classification"] == "requires_human_review")
    return len(duplicates) + empty_review + len(non_pd)


def write_rows(sheet, rows: list[dict[str, Any]], columns: list[str]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])


def style_sheet(sheet, decision_values: list[str] | None = None) -> None:
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    input_columns = []
    for cell in sheet[1]:
        if cell.value in {"human_decision", "human_notes"}:
            input_columns.append(cell.column)
    for column_idx in input_columns:
        letter = get_column_letter(column_idx)
        for cell in sheet[letter][1:]:
            cell.fill = INPUT_FILL
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)
    if decision_values:
        human_decision_col = None
        for cell in sheet[1]:
            if cell.value == "human_decision":
                human_decision_col = cell.column
                break
        if human_decision_col:
            letter = get_column_letter(human_decision_col)
            formula = '"' + ",".join(decision_values) + '"'
            validation = DataValidation(type="list", formula1=formula, allow_blank=True)
            validation.error = "Choose one of the suggested review decisions."
            validation.errorTitle = "Invalid decision"
            sheet.add_data_validation(validation)
            validation.add(f"{letter}2:{letter}{max(sheet.max_row, 2)}")


def add_summary_sheet(workbook: Workbook, stats: dict[str, Any], duplicate_rows_count: int, review_count: int) -> None:
    sheet = workbook.create_sheet("review_summary")
    rows = [
        ("metric", "value"),
        ("total_product_rows", stats["total_products"]),
        ("filled_part_nr_sku", stats["part_nr_sku_filled"]),
        ("empty_part_nr_sku", stats["part_nr_sku_empty"]),
        ("pd_pattern_rows", stats["part_nr_sku_pd_pattern_count"]),
        ("non_pd_pattern_rows", stats["part_nr_sku_non_pd_pattern_count"]),
        ("duplicate_groups", stats["duplicate_group_count"]),
        ("duplicate_rows", duplicate_rows_count),
        ("rows_requiring_human_review", review_count),
        ("product_status", "manually_confirmed_pending_duplicate_validation"),
    ]
    for row in rows:
        sheet.append(row)
    style_sheet(sheet)


def add_decision_log_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("decision_log")
    sheet.append(["decision_date", "reviewer", "topic", "decision", "rationale", "follow_up_action"])
    style_sheet(sheet)


def build_workbook(path: Path, duplicate_data: list[dict[str, Any]], empty_data: list[dict[str, Any]], non_pd_data: list[dict[str, Any]], stats: dict[str, Any]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    duplicate_columns = list(duplicate_data[0].keys()) if duplicate_data else [
        "duplicate_group_id",
        "source_row_number",
        "part_nr_sku",
        "pd_ref_nr",
        "classification_from_audit",
        "suggested_review_question",
        "human_decision",
        "human_notes",
    ]
    duplicate_sheet = workbook.create_sheet("duplicate_part_nr_sku")
    write_rows(duplicate_sheet, duplicate_data, duplicate_columns)
    style_sheet(duplicate_sheet, DUPLICATE_DECISIONS)

    empty_columns = list(empty_data[0].keys()) if empty_data else [
        "source_row_number",
        "pd_ref_nr",
        "current_audit_classification",
        "suggested_action",
        "human_decision",
        "human_notes",
    ]
    empty_sheet = workbook.create_sheet("empty_part_nr_sku")
    write_rows(empty_sheet, empty_data, empty_columns)
    style_sheet(empty_sheet, EMPTY_DECISIONS)

    non_pd_columns = list(non_pd_data[0].keys()) if non_pd_data else [
        "source_row_number",
        "part_nr_sku",
        "pd_ref_nr",
        "suggested_interpretation",
        "human_decision",
        "human_notes",
    ]
    non_pd_sheet = workbook.create_sheet("non_pd_pattern_products")
    write_rows(non_pd_sheet, non_pd_data, non_pd_columns)
    style_sheet(non_pd_sheet, NON_PD_DECISIONS)

    review_count = rows_requiring_human_review(duplicate_data, empty_data, non_pd_data)
    add_summary_sheet(workbook, stats, len(duplicate_data), review_count)
    add_decision_log_sheet(workbook)
    workbook.save(path)


def run_product_reference_review_spreadsheet(
    data_dir: Path = DATA_DIR,
    output_dir: Path = OUTPUT_DIR,
) -> ProductReferenceReviewSpreadsheetResult:
    df = product_dataframe(data_dir)
    columns = relevant_product_columns(df)
    duplicate_audits = duplicate_group_audit(df)
    empty_audits = empty_reference_audit(df)
    duplicate_data = duplicate_rows(df, duplicate_audits, columns)
    empty_data = empty_rows(df, empty_audits, columns)
    non_pd_data = non_pd_rows(df, columns)
    stats = audit_stats(df, duplicate_audits, empty_audits)

    ensure_dir(output_dir)
    xlsx_path = output_dir / "product_reference_human_review.xlsx"
    backup_existing(xlsx_path, run_id())
    build_workbook(xlsx_path, duplicate_data, empty_data, non_pd_data, stats)

    return ProductReferenceReviewSpreadsheetResult(
        output_dir=output_dir,
        xlsx_path=xlsx_path,
        duplicate_rows=len(duplicate_data),
        empty_rows=len(empty_data),
        non_pd_rows=len(non_pd_data),
        rows_requiring_human_review=rows_requiring_human_review(duplicate_data, empty_data, non_pd_data),
    )
