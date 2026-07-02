from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .source_onboarding import backup_existing, ensure_dir


STEP3D_DIR = Path("outputs/originaldatabase_analysis/step3d_serial_aware_review")
OUTPUT_DIR = Path("outputs/originaldatabase_analysis/step3e_human_approval_spreadsheet")
CONFIG_DIR = Path("config/data_model")

DECISION_OPTIONS = [
    ("approved", "Approve the candidate exactly as reviewed."),
    ("approved_as_semantic_primary_key", "Approve a semantic primary key based on uniqueness and serial evidence."),
    ("approved_as_header_line_relationship", "Approve a header-line relationship based on match rate and join risk."),
    ("approved_as_technical_key_only", "Approve only as an analytical technical key, not as an ERP business key."),
    ("rejected", "Reject the candidate."),
    ("needs_business_context", "Hold until business meaning or grain is clarified."),
    ("do_not_approve_yet", "Do not approve at this stage because technical risk is unresolved."),
    ("pending", "No human decision has been made yet."),
]

MATRIX_COLUMNS = [
    "decision_id",
    "decision_group",
    "decision_type",
    "table_name",
    "source_table",
    "source_column",
    "target_table",
    "target_column",
    "candidate",
    "semantic_namespace",
    "semantic_ref_name",
    "expected_prefix",
    "evidence_summary",
    "non_null_rate",
    "uniqueness_rate",
    "prefix_match_rate",
    "regex_match_rate",
    "match_rate",
    "join_risk",
    "combined_confidence",
    "recommended_human_decision",
    "current_status",
    "risk_explanation",
    "business_question",
    "human_decision",
    "human_notes",
]


@dataclass(frozen=True)
class ApprovalSpreadsheetResult:
    output_dir: Path
    xlsx_path: Path
    csv_path: Path
    decision_count: int
    primary_key_count: int
    relationship_count: int
    technical_key_count: int
    needs_context_count: int
    conflict_count: int


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def write_csv_rows(path: Path, rows: list[dict[str, Any]], current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=MATRIX_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def safe_id(value: str) -> str:
    text = re.sub(r"[^A-Za-z0-9]+", "_", value).strip("_").upper()
    return text[:80] or "DECISION"


def key_group(row: dict[str, str]) -> str:
    if row["recommended_human_decision"] == "approve_as_semantic_primary_key":
        return "semantic primary key"
    if row["recommended_human_decision"] == "approve_as_technical_key_only":
        return "technical key"
    if row["status"] == "needs_business_context":
        return "needs business context"
    return "other key candidate"


def relationship_group(row: dict[str, str]) -> str:
    if row["recommended_human_decision"] == "approve_header_line_relationship":
        return "header-line relationship"
    if row["status"] == "needs_business_context":
        return "needs business context"
    if row["status"] == "rejected_proposal":
        return "conflict"
    return "other relationship candidate"


def risk_explanation_for_key(row: dict[str, str]) -> str:
    if row["recommended_human_decision"] == "approve_as_semantic_primary_key":
        return "Strong technical evidence: uniqueness, non-null, prefix, and regex all align."
    if row["recommended_human_decision"] == "approve_as_technical_key_only":
        return "Line table key is suitable as a technical analytical key, not an ERP business key."
    if row["status"] == "needs_business_context":
        return "Technical evidence is insufficient or table meaning/grain needs confirmation."
    return "Candidate requires manual review before use."


def risk_explanation_for_relationship(row: dict[str, str]) -> str:
    if row["recommended_human_decision"] == "approve_header_line_relationship":
        return "Header-line candidate has full match, low join risk, and serial prefix consistency."
    if row["join_risk"] == "high":
        return "High join risk or target-side duplicates require review before approval."
    if row["status"] == "rejected_proposal":
        return "Technical measurements do not support approval at this stage."
    return "Relationship requires manual review before use."


def business_question_for_key(row: dict[str, str]) -> str:
    table = row["table_name"]
    candidate = row["candidate_key"]
    if table == "product_export_product" and candidate in {"part_nr_sku", "product_code", "sku", "item_code"}:
        return "Is this the principal product code or only an internal SKU/reference?"
    if table == "organisation_export_organisation":
        return "Does Organisation use its own serial reference, or is it a master-data record without serial semantics?"
    if row["recommended_human_decision"] == "approve_as_technical_key_only":
        return "Approve this only as an analytical technical key for line-level work?"
    if row["status"] == "needs_business_context":
        return "What is the approved business grain and semantic meaning for this candidate?"
    return "Approve this recommendation, reject it, or request more context?"


def business_question_for_relationship(row: dict[str, str]) -> str:
    if row["source_table"] == "goodsreception_export_goodsreception" and row["source_column"] == "ref_nr_purchase_order":
        return "Does this field link Goods Reception with Purchase Order?"
    if row["join_risk"] == "high":
        return "Can this relationship create duplicate joins or does it require a different grain?"
    if row["recommended_human_decision"] == "approve_header_line_relationship":
        return "Approve this as a header-line relationship?"
    return "Approve, reject, or request business context for this relationship?"


def build_key_decision(row: dict[str, str], index: int) -> dict[str, Any]:
    decision_group = key_group(row)
    decision_type = "primary_key" if decision_group == "semantic primary key" else "technical_key" if decision_group == "technical key" else "needs_business_context"
    return {
        "decision_id": f"KEY_{index:03d}_{safe_id(row['table_name'] + '_' + row['candidate_key'])}",
        "decision_group": decision_group,
        "decision_type": decision_type,
        "table_name": row["table_name"],
        "source_table": "",
        "source_column": "",
        "target_table": "",
        "target_column": "",
        "candidate": row["candidate_key"],
        "semantic_namespace": row["semantic_namespace"],
        "semantic_ref_name": row["semantic_ref_name"],
        "expected_prefix": row["expected_prefix"],
        "evidence_summary": f"non_null={row['non_null_rate']}%; uniqueness={row['uniqueness_rate']}%; prefix={row['prefix_match_rate']}%; regex={row['regex_match_rate']}%",
        "non_null_rate": row["non_null_rate"],
        "uniqueness_rate": row["uniqueness_rate"],
        "prefix_match_rate": row["prefix_match_rate"],
        "regex_match_rate": row["regex_match_rate"],
        "match_rate": "",
        "join_risk": "",
        "combined_confidence": row["combined_confidence"],
        "recommended_human_decision": row["recommended_human_decision"],
        "current_status": row["status"],
        "risk_explanation": risk_explanation_for_key(row),
        "business_question": business_question_for_key(row),
        "human_decision": "pending",
        "human_notes": "",
    }


def build_relationship_decision(row: dict[str, str], index: int) -> dict[str, Any]:
    decision_group = relationship_group(row)
    decision_type = "relationship" if decision_group == "header-line relationship" else "needs_business_context"
    return {
        "decision_id": f"REL_{index:03d}_{safe_id(row['source_table'] + '_' + row['target_table'])}",
        "decision_group": decision_group,
        "decision_type": decision_type,
        "table_name": "",
        "source_table": row["source_table"],
        "source_column": row["source_column"],
        "target_table": row["target_table"],
        "target_column": row["target_column"],
        "candidate": f"{row['source_table']}.{row['source_column']} -> {row['target_table']}.{row['target_column']}",
        "semantic_namespace": f"{row['source_semantic_namespace']} -> {row['target_semantic_namespace']}",
        "semantic_ref_name": "",
        "expected_prefix": row["expected_prefix"],
        "evidence_summary": f"match={row['match_rate']}%; unmatched={row['unmatched_count']}; target_duplicates={row['target_duplicate_count']}; join_risk={row['join_risk']}; prefix={row['prefix_consistency']}",
        "non_null_rate": "",
        "uniqueness_rate": "",
        "prefix_match_rate": "",
        "regex_match_rate": "",
        "match_rate": row["match_rate"],
        "join_risk": row["join_risk"],
        "combined_confidence": row["combined_confidence"],
        "recommended_human_decision": row["recommended_human_decision"],
        "current_status": row["status"],
        "risk_explanation": risk_explanation_for_relationship(row),
        "business_question": business_question_for_relationship(row),
        "human_decision": "pending",
        "human_notes": "",
    }


def parse_conflicts(conflict_text: str) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    blocks = re.split(r"\n## ", conflict_text)
    for block in blocks[1:]:
        lines = block.splitlines()
        table = lines[0].strip()
        data = {"table": table}
        for line in lines:
            if line.startswith("- Source file:"):
                data["source_file"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Row count:"):
                data["row_count"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Expected prefix:"):
                data["expected_prefix"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Detected prefix:"):
                data["detected_prefix"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Technical hypothesis:"):
                data["technical_hypothesis"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Risk:"):
                data["risk"] = line.split(":", 1)[1].strip()
            elif line.startswith("- Question for human review:"):
                data["question"] = line.split(":", 1)[1].strip()
        rows.append(
            {
                "decision_id": f"CONFLICT_{safe_id(table)}",
                "decision_group": "conflict",
                "decision_type": "needs_business_context",
                "table_name": table,
                "source_table": "",
                "source_column": "",
                "target_table": "",
                "target_column": "",
                "candidate": table,
                "semantic_namespace": "",
                "semantic_ref_name": "",
                "expected_prefix": data.get("expected_prefix", ""),
                "evidence_summary": f"detected_prefix={data.get('detected_prefix', '')}; rows={data.get('row_count', '')}",
                "non_null_rate": "",
                "uniqueness_rate": "",
                "prefix_match_rate": "",
                "regex_match_rate": "",
                "match_rate": "",
                "join_risk": "high",
                "combined_confidence": "needs_business_context",
                "recommended_human_decision": "do_not_approve_yet",
                "current_status": "needs_business_context",
                "risk_explanation": data.get("risk", "Conflict requires review before approval."),
                "business_question": data.get("question", "Resolve the table classification before approving."),
                "human_decision": "pending",
                "human_notes": "",
            }
        )
    return rows


def build_decisions(step3d_dir: Path) -> tuple[list[dict[str, Any]], dict[str, list[dict[str, Any]]]]:
    key_rows = read_csv_rows(step3d_dir / "serial_aware_key_review.csv")
    relationship_rows = read_csv_rows(step3d_dir / "serial_aware_relationship_review.csv")
    conflict_text = (step3d_dir / "conflict_investigation.md").read_text(encoding="utf-8")

    primary_keys = [
        build_key_decision(row, idx)
        for idx, row in enumerate(key_rows, 1)
        if row["recommended_human_decision"] == "approve_as_semantic_primary_key"
    ]
    technical_keys = [
        build_key_decision(row, idx)
        for idx, row in enumerate(key_rows, 1)
        if row["recommended_human_decision"] == "approve_as_technical_key_only"
    ]
    relationships = [
        build_relationship_decision(row, idx)
        for idx, row in enumerate(relationship_rows, 1)
        if row["recommended_human_decision"] == "approve_header_line_relationship"
    ]
    needs_context = [
        build_key_decision(row, idx)
        for idx, row in enumerate(key_rows, 1)
        if row["status"] == "needs_business_context"
    ] + [
        build_relationship_decision(row, idx)
        for idx, row in enumerate(relationship_rows, 1)
        if row["status"] == "needs_business_context"
    ]
    conflicts = parse_conflicts(conflict_text)

    all_rows = primary_keys + relationships + technical_keys + needs_context + conflicts
    grouped = {
        "Primary Keys": primary_keys,
        "Relationships": relationships,
        "Technical Line Keys": technical_keys,
        "Needs Context": needs_context,
        "Conflicts": conflicts,
    }
    return all_rows, grouped


def add_rows(sheet, rows: list[list[Any]]) -> None:
    for row in rows:
        sheet.append(row)


def write_table_sheet(sheet, rows: list[dict[str, Any]], columns: list[str]) -> None:
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])


def style_sheet(sheet, human_decision_column: str | None = None) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 45)
    if human_decision_column:
        idx = None
        for cell in sheet[1]:
            if cell.value == human_decision_column:
                idx = cell.column
                break
        if idx:
            letter = get_column_letter(idx)
            fill = PatternFill("solid", fgColor="FFF2CC")
            for cell in sheet[letter]:
                cell.fill = fill if cell.row > 1 else PatternFill("solid", fgColor="9E6A03")
            sheet.column_dimensions[letter].width = 28


def add_decision_validation(workbook: Workbook, sheet_names: list[str]) -> None:
    formula = "'Decision Options'!$A$2:$A$9"
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        human_idx = None
        for cell in sheet[1]:
            if cell.value == "human_decision":
                human_idx = cell.column
                break
        if not human_idx:
            continue
        letter = get_column_letter(human_idx)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Choose one of the allowed decision options."
        dv.errorTitle = "Invalid decision"
        sheet.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}{max(sheet.max_row, 2)}")


def apply_group_colors(sheet) -> None:
    colors = {
        "semantic primary key": "E2F0D9",
        "header-line relationship": "DDEBF7",
        "technical key": "FCE4D6",
        "needs business context": "FFF2CC",
        "conflict": "F4CCCC",
    }
    group_col = None
    for cell in sheet[1]:
        if cell.value == "decision_group":
            group_col = cell.column
            break
    if not group_col:
        return
    for row in range(2, sheet.max_row + 1):
        value = sheet.cell(row, group_col).value
        color = colors.get(value)
        if not color:
            continue
        for col in range(1, sheet.max_column + 1):
            sheet.cell(row, col).fill = PatternFill("solid", fgColor=color)


def render_readme_md(path: Path, current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    path.write_text(
        "# Human Approval Matrix\n\n"
        "This spreadsheet is for human review only. No key or relationship is approved automatically.\n\n"
        "Use the `Approval Matrix` sheet as the main review queue. Fill `human_decision` using the allowed values in `Decision Options`, and add context in `human_notes` when needed.\n\n"
        "The approval application step is intentionally not run here. A future controlled command can apply only manually filled decisions.\n",
        encoding="utf-8",
    )


def build_workbook(xlsx_path: Path, decisions: list[dict[str, Any]], grouped: dict[str, list[dict[str, Any]]], current_run_id: str) -> None:
    ensure_dir(xlsx_path.parent)
    backup_existing(xlsx_path, current_run_id)
    workbook = Workbook()
    workbook.remove(workbook.active)

    readme = workbook.create_sheet("README")
    add_rows(
        readme,
        [
            ["Human Approval Spreadsheet"],
            ["This workbook is only for human review. No decision has been applied automatically."],
            ["Fill the human_decision column. Allowed values are listed in Decision Options."],
            ["A later controlled step can apply only manually completed decisions."],
        ],
    )
    readme["A1"].font = Font(bold=True, size=16)
    readme.column_dimensions["A"].width = 120

    matrix = workbook.create_sheet("Approval Matrix")
    write_table_sheet(matrix, decisions, MATRIX_COLUMNS)

    for sheet_name in ["Primary Keys", "Relationships", "Technical Line Keys", "Needs Context", "Conflicts"]:
        sheet = workbook.create_sheet(sheet_name)
        write_table_sheet(sheet, grouped[sheet_name], MATRIX_COLUMNS)

    options = workbook.create_sheet("Decision Options")
    options.append(["human_decision", "description"])
    for option, description in DECISION_OPTIONS:
        options.append([option, description])

    for sheet in workbook.worksheets:
        style_sheet(sheet, "human_decision" if sheet.title != "Decision Options" and sheet.title != "README" else None)
        if sheet.title in {"Approval Matrix", "Primary Keys", "Relationships", "Technical Line Keys", "Needs Context", "Conflicts"}:
            apply_group_colors(sheet)
            # Re-highlight human_decision after row group coloring.
            for cell in sheet[1]:
                if cell.value == "human_decision":
                    letter = get_column_letter(cell.column)
                    for decision_cell in sheet[letter][1:]:
                        decision_cell.fill = PatternFill("solid", fgColor="FFF2CC")
                    break
    add_decision_validation(
        workbook,
        ["Approval Matrix", "Primary Keys", "Relationships", "Technical Line Keys", "Needs Context", "Conflicts"],
    )
    workbook.save(xlsx_path)


def verify_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    expected = {
        "README",
        "Approval Matrix",
        "Primary Keys",
        "Relationships",
        "Technical Line Keys",
        "Needs Context",
        "Conflicts",
        "Decision Options",
    }
    missing = expected - set(workbook.sheetnames)
    if missing:
        raise ValueError(f"Missing workbook sheets: {sorted(missing)}")
    matrix = workbook["Approval Matrix"]
    headers = [cell.value for cell in matrix[1]]
    if "human_decision" not in headers:
        raise ValueError("Approval Matrix is missing human_decision.")


def run_approval_spreadsheet(
    step3d_dir: Path = STEP3D_DIR,
    output_dir: Path = OUTPUT_DIR,
    config_dir: Path = CONFIG_DIR,
) -> ApprovalSpreadsheetResult:
    current_run_id = run_id()
    decisions, grouped = build_decisions(step3d_dir)
    xlsx_path = output_dir / "human_approval_matrix.xlsx"
    csv_path = output_dir / "human_approval_matrix.csv"
    readme_path = output_dir / "human_approval_matrix_readme.md"

    write_csv_rows(csv_path, decisions, current_run_id)
    build_workbook(xlsx_path, decisions, grouped, current_run_id)
    render_readme_md(readme_path, current_run_id)
    verify_workbook(xlsx_path)

    return ApprovalSpreadsheetResult(
        output_dir=output_dir,
        xlsx_path=xlsx_path,
        csv_path=csv_path,
        decision_count=len(decisions),
        primary_key_count=len(grouped["Primary Keys"]),
        relationship_count=len(grouped["Relationships"]),
        technical_key_count=len(grouped["Technical Line Keys"]),
        needs_context_count=len(grouped["Needs Context"]),
        conflict_count=len(grouped["Conflicts"]),
    )
