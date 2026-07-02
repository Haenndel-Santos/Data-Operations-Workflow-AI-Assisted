from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .product_reference_audit import OUTPUT_DIR, clean_value
from .source_onboarding import backup_existing, ensure_dir


WORKBOOK_NAME = "product_reference_human_review.xlsx"
REPORT_NAME = "product_reference_final_decision.md"
REVIEW_SHEETS = ["duplicate_part_nr_sku", "empty_part_nr_sku", "non_pd_pattern_products"]
DECISION_LOG_SHEET = "decision_log"

UNRESOLVED_DECISIONS = {"", "pending", "requires_more_investigation"}
MORE_INVESTIGATION = "requires_more_investigation"
VALID_TEXTUAL = "valid_textual_product_reference"
DUPLICATE_RECORD = "same_product_duplicate_record"
DISTINCT_SAME_REFERENCE = "distinct_products_same_reference"
EXCLUDE_DECISIONS = {"exclude_from_product_master", "obsolete_or_inactive_record"}
REPAIR_DECISION = "repair_reference"


@dataclass(frozen=True)
class ProductReferenceFinalDecisionResult:
    output_dir: Path
    workbook_path: Path
    report_path: Path
    total_reviewed_rows: int
    unresolved_rows: int
    more_investigation_rows: int
    distinct_same_reference_rows: int
    part_nr_sku_unique_key_recommended: bool


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def rows_from_sheet(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_value(value) for value in rows[0]]
    output = []
    for raw_row in rows[1:]:
        if raw_row is None or not any(clean_value(value) for value in raw_row):
            continue
        output.append({header: clean_value(value) for header, value in zip(headers, raw_row, strict=False) if header})
    return output


def decision_value(row: dict[str, Any]) -> str:
    return clean_value(row.get("human_decision", "")).casefold()


def note_value(row: dict[str, Any]) -> str:
    return clean_value(row.get("human_notes", ""))


def row_identifier(row: dict[str, Any], fallback: int) -> str:
    for key in ("source_row_number", "duplicate_group_id", "part_nr_sku"):
        value = clean_value(row.get(key, ""))
        if value:
            return f"{key}={value}"
    return f"row_index={fallback}"


def decision_log_entries(workbook) -> list[dict[str, Any]]:
    return rows_from_sheet(workbook, DECISION_LOG_SHEET)


def summarize(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    sheet_counts = {sheet: len(rows) for sheet, rows in rows_by_sheet.items()}
    decision_counts: Counter[str] = Counter()
    notes_by_sheet: dict[str, list[dict[str, str]]] = defaultdict(list)
    unresolved: list[dict[str, str]] = []
    more_investigation: list[dict[str, str]] = []
    valid_textual: list[dict[str, str]] = []
    duplicate_records: list[dict[str, str]] = []
    distinct_same_reference: list[dict[str, str]] = []
    excluded: list[dict[str, str]] = []
    repair: list[dict[str, str]] = []

    for sheet_name, rows in rows_by_sheet.items():
        for index, row in enumerate(rows, start=2):
            decision = decision_value(row)
            note = note_value(row)
            identifier = row_identifier(row, index)
            decision_counts[decision or "unresolved"] += 1
            entry = {"sheet": sheet_name, "identifier": identifier, "decision": decision or "unresolved", "notes": note}
            if note:
                notes_by_sheet[sheet_name].append(entry)
            if decision in UNRESOLVED_DECISIONS:
                unresolved.append(entry)
            if decision == MORE_INVESTIGATION:
                more_investigation.append(entry)
            if decision == VALID_TEXTUAL:
                valid_textual.append(entry)
            if decision == DUPLICATE_RECORD:
                duplicate_records.append(entry)
            if decision == DISTINCT_SAME_REFERENCE:
                distinct_same_reference.append(entry)
            if decision in EXCLUDE_DECISIONS:
                excluded.append(entry)
            if decision == REPAIR_DECISION:
                repair.append(entry)

    return {
        "sheet_counts": sheet_counts,
        "decision_counts": dict(sorted(decision_counts.items())),
        "notes_by_sheet": dict(notes_by_sheet),
        "unresolved": unresolved,
        "more_investigation": more_investigation,
        "valid_textual": valid_textual,
        "duplicate_records": duplicate_records,
        "distinct_same_reference": distinct_same_reference,
        "excluded": excluded,
        "repair": repair,
        "total_reviewed_rows": sum(sheet_counts.values()),
    }


def can_use_part_nr_sku_as_unique_business_constraint(summary: dict[str, Any]) -> bool:
    return (
        not summary["distinct_same_reference"]
        and not summary["repair"]
        and not summary["unresolved"]
        and not summary["more_investigation"]
    )


def list_entries(entries: list[dict[str, str]], empty_text: str, limit: int = 25) -> list[str]:
    if not entries:
        return [f"- {empty_text}"]
    lines = []
    for entry in entries[:limit]:
        note = f"; notes={entry['notes']}" if entry.get("notes") else ""
        lines.append(f"- {entry['sheet']} / {entry['identifier']}: decision={entry['decision']}{note}")
    remaining = len(entries) - limit
    if remaining > 0:
        lines.append(f"- ... {remaining} more rows not shown")
    return lines


def render_report(summary: dict[str, Any], decision_log: list[dict[str, Any]]) -> str:
    unique_business_constraint_ok = can_use_part_nr_sku_as_unique_business_constraint(summary)
    unique_business_statement = (
        "`part_nr_sku` can be considered for a unique business constraint only after exclusions/cleanup are applied and revalidated."
        if unique_business_constraint_ok
        else "`part_nr_sku` should not be treated as unique until the open decisions and cleanup items are resolved."
    )
    business_reference_statement = (
        "`part_nr_sku` should remain the main business reference/search/matching field."
    )
    primary_key_statement = "`part_nr_sku` must not be approved as the Product primary key; use generated `product_id` instead."

    lines = [
        "# Product Reference Final Decision",
        "",
        "This report consolidates manually entered decisions from `product_reference_human_review.xlsx`. The workbook is read only by this step.",
        "",
        "## Reviewed Rows By Sheet",
        "",
    ]
    for sheet, count in summary["sheet_counts"].items():
        lines.append(f"- `{sheet}`: {count}")

    lines.extend(["", "## Decision Counts", ""])
    for decision, count in summary["decision_counts"].items():
        lines.append(f"- `{decision}`: {count}")

    lines.extend(
        [
            "",
            "## Human Notes",
            "",
        ]
    )
    if not summary["notes_by_sheet"]:
        lines.append("- No human notes were provided.")
    else:
        for sheet_name, entries in summary["notes_by_sheet"].items():
            lines.append(f"### {sheet_name}")
            lines.extend(list_entries(entries, "No notes.", limit=20))

    sections = [
        ("Unresolved Issues", summary["unresolved"], "No unresolved rows."),
        ("Rows Requiring More Investigation", summary["more_investigation"], "No rows marked as requiring more investigation."),
        ("Valid Textual Product References", summary["valid_textual"], "No rows marked as valid textual product references."),
        ("Duplicate Records", summary["duplicate_records"], "No rows marked as same-product duplicate records."),
        ("Distinct Products Sharing Reference", summary["distinct_same_reference"], "No rows marked as distinct products sharing the same reference."),
        ("Rows To Exclude", summary["excluded"], "No rows marked for exclusion."),
        ("Rows Requiring Reference Repair", summary["repair"], "No rows marked for reference repair."),
    ]
    for title, entries, empty_text in sections:
        lines.extend(["", f"## {title}", ""])
        lines.extend(list_entries(entries, empty_text))

    lines.extend(
        [
            "",
            "## Final Product Key Recommendation",
            "",
            "- Can `part_nr_sku` be used as the Product primary key? No",
            f"- Primary-key rationale: {primary_key_statement}",
            f"- Can `part_nr_sku` be considered as a unique business constraint after cleanup? {'Yes' if unique_business_constraint_ok else 'No'}",
            f"- Unique-business-constraint rationale: {unique_business_statement}",
            "- Should `part_nr_sku` remain only as a business reference/search/matching field in the final model? Yes",
            f"- Business-reference recommendation: {business_reference_statement}",
            "- `pd_ref_nr` should remain optional and populated only when `part_nr_sku` matches the serial-style PD pattern.",
            "- Do not update `approved_keys.yml` or `approved_relationships.yml` from this report.",
            "",
            "## Recommended Final Product Model",
            "",
            "- `product_id`: generated technical primary key",
            "- `part_nr_sku`: main business reference/search/matching field",
            "- `pd_ref_nr`: optional serial-style reference where available",
            "- `source_system`: original source",
            "- `source_record_id`: original row/record identifier",
            "- `review_status`: validation status",
            "- `is_active`: active/inactive status, if available",
            "",
            "## Required Cleanup Actions Before Migration/Import",
            "",
            "- Resolve all rows with empty or unresolved `human_decision`.",
            "- Resolve every row marked `requires_more_investigation`.",
            "- Repair rows marked `repair_reference` before loading into the target model.",
            "- Exclude or archive rows marked for exclusion before uniqueness checks.",
            "- If any row is confirmed as `distinct_products_same_reference`, keep `product_id` as the only global primary key and enforce `part_nr_sku` as non-unique business reference.",
            "- Even when cleanup makes `part_nr_sku` unique, keep `product_id` as the technical primary key for import/migration stability.",
            "- Re-run duplicate validation after cleanup and before any migration/import.",
        ]
    )

    lines.extend(["", "## Decision Log", ""])
    if not decision_log:
        lines.append("- No decision log entries were provided.")
    else:
        for entry in decision_log:
            lines.append(
                "- date={decision_date}; reviewer={reviewer}; topic={topic}; decision={decision}; follow_up={follow_up_action}".format(
                    decision_date=clean_value(entry.get("decision_date", "")) or "not provided",
                    reviewer=clean_value(entry.get("reviewer", "")) or "not provided",
                    topic=clean_value(entry.get("topic", "")) or "not provided",
                    decision=clean_value(entry.get("decision", "")) or "not provided",
                    follow_up_action=clean_value(entry.get("follow_up_action", "")) or "not provided",
                )
            )
            rationale = clean_value(entry.get("rationale", ""))
            if rationale:
                lines.append(f"  Rationale: {rationale}")

    return "\n".join(lines)


def run_product_reference_final_decision(
    output_dir: Path = OUTPUT_DIR,
    workbook_path: Path | None = None,
) -> ProductReferenceFinalDecisionResult:
    workbook_path = workbook_path or output_dir / WORKBOOK_NAME
    if not workbook_path.exists():
        raise FileNotFoundError(f"Product human review workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True)
    rows_by_sheet = {sheet: rows_from_sheet(workbook, sheet) for sheet in REVIEW_SHEETS}
    log_entries = decision_log_entries(workbook)
    summary = summarize(rows_by_sheet)

    ensure_dir(output_dir)
    report_path = output_dir / REPORT_NAME
    backup_existing(report_path, run_id())
    report_path.write_text(render_report(summary, log_entries), encoding="utf-8")

    return ProductReferenceFinalDecisionResult(
        output_dir=output_dir,
        workbook_path=workbook_path,
        report_path=report_path,
        total_reviewed_rows=summary["total_reviewed_rows"],
        unresolved_rows=len(summary["unresolved"]),
        more_investigation_rows=len(summary["more_investigation"]),
        distinct_same_reference_rows=len(summary["distinct_same_reference"]),
        part_nr_sku_unique_key_recommended=False,
    )
