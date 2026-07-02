from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .product_reference_audit import OUTPUT_DIR, clean_value
from .product_refnr_final_review_spreadsheet import XLSX_NAME
from .product_refnr_final_review_validation import (
    REPORT_NAME,
    SUMMARY_NAME,
    consolidated_rows,
    missing_final_note,
    normalized_final_decision,
)
from .product_refnr_human_review import DECISION_OPTIONS
from .source_onboarding import backup_existing, ensure_dir


XLSX_OUTPUT_NAME = "product_refnr_missing_notes_fix.xlsx"
CSV_OUTPUT_NAME = "product_refnr_missing_notes_fix.csv"
README_OUTPUT_NAME = "product_refnr_missing_notes_fix_readme.md"

NOTE_SUGGESTIONS = {
    "approved_use_corrected_product_ref_nr": "Corrected product reference accepted after manual review.",
    "approved_keep_original_part_nr_sku_only": "No corrected Product_ref.nr match found; original part_nr_sku kept as functional product reference.",
    "approved_create_technical_product_id_only": "No reliable natural product reference confirmed; product retained with generated technical product_id only.",
    "merge_duplicate_records": "Duplicate records confirmed as the same product after manual review.",
    "keep_as_separate_products": "Records confirmed as separate products after manual review.",
    "needs_business_context": "Requires additional business context before final product modeling decision.",
    "rejected": "Record rejected after manual review.",
}

FIX_COLUMNS = [
    "fix_id",
    "original_sheet",
    "original_excel_row",
    "issue_id",
    "issue_type",
    "product_original_identifier",
    "product_refnr_identifier",
    "original_part_nr_sku",
    "corrected_product_ref_nr",
    "current_final_human_decision",
    "current_final_human_notes",
    "suggested_final_human_notes",
    "final_human_notes_to_apply",
]


@dataclass(frozen=True)
class ProductRefnrMissingNotesFixResult:
    output_dir: Path
    xlsx_path: Path
    csv_path: Path
    readme_path: Path
    missing_notes_count: int


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def suggested_note(decision: str) -> str:
    return NOTE_SUGGESTIONS.get(decision, "")


def fix_row(row: dict[str, Any], index: int) -> dict[str, Any]:
    decision = normalized_final_decision(row)
    suggestion = suggested_note(decision)
    return {
        "fix_id": f"MISSING_NOTE_{index:03d}",
        "original_sheet": clean_value(row.get("original_sheet", "")),
        "original_excel_row": clean_value(row.get("original_excel_row", "")),
        "issue_id": clean_value(row.get("issue_id", "")),
        "issue_type": clean_value(row.get("issue_type", "")),
        "product_original_identifier": clean_value(row.get("product_original_identifier", "")),
        "product_refnr_identifier": clean_value(row.get("product_refnr_identifier", "")),
        "original_part_nr_sku": clean_value(row.get("original_part_nr_sku", "")),
        "corrected_product_ref_nr": clean_value(row.get("corrected_product_ref_nr", "")),
        "current_final_human_decision": decision,
        "current_final_human_notes": clean_value(row.get("final_human_notes", "")),
        "suggested_final_human_notes": suggestion,
        "final_human_notes_to_apply": suggestion,
    }


def collect_missing_note_rows(workbook) -> list[dict[str, Any]]:
    rows = [row for row in consolidated_rows(workbook) if missing_final_note(row)]
    return [fix_row(row, index) for index, row in enumerate(rows, start=1)]


def style_sheet(sheet, input_columns: set[str] | None = None) -> None:
    input_columns = input_columns or set()
    header_fill = PatternFill("solid", fgColor="1F4E78")
    input_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    headers = [cell.value for cell in sheet[1]]
    input_indexes = {headers.index(column) + 1 for column in input_columns if column in headers}
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if cell.column in input_indexes:
                cell.fill = input_fill
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 14), 72)


def write_readme_sheet(workbook: Workbook, missing_count: int) -> None:
    sheet = workbook.create_sheet("README")
    rows = [
        ["Product RefNr Missing Notes Fix"],
        ["This workbook is an auxiliary copy/paste aid only."],
        ["It contains only rows with missing final_human_notes from the Product final review workbook."],
        ["No Product decisions are applied automatically by this file."],
        [f"Missing notes included: {missing_count}"],
        ["Copy the suggested note into the original final review workbook, or replace it with a more specific business note."],
    ]
    for row in rows:
        sheet.append(row)
    sheet["A1"].font = Font(bold=True, size=16)
    sheet.column_dimensions["A"].width = 130


def write_missing_notes_sheet(workbook: Workbook, rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet("Missing Notes Fix")
    sheet.append(FIX_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in FIX_COLUMNS])
    style_sheet(sheet, {"final_human_notes_to_apply"})


def write_suggested_notes_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Suggested Notes")
    sheet.append(["final_human_decision", "suggested_final_human_notes"])
    for decision, note in NOTE_SUGGESTIONS.items():
        sheet.append([decision, note])
    style_sheet(sheet)


def write_decision_options_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("Decision Options")
    sheet.append(["final_human_decision"])
    for option in DECISION_OPTIONS:
        sheet.append([option])
    style_sheet(sheet)


def write_workbook(path: Path, rows: list[dict[str, Any]], current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    workbook = Workbook()
    workbook.remove(workbook.active)
    write_readme_sheet(workbook, len(rows))
    write_missing_notes_sheet(workbook, rows)
    write_suggested_notes_sheet(workbook)
    write_decision_options_sheet(workbook)
    workbook.save(path)


def write_csv(path: Path, rows: list[dict[str, Any]], current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=FIX_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def render_readme(rows: list[dict[str, Any]], workbook_path: Path, validation_report: Path, validation_summary: Path) -> str:
    by_decision: dict[str, int] = {}
    for row in rows:
        decision = clean_value(row.get("current_final_human_decision", ""))
        by_decision[decision] = by_decision.get(decision, 0) + 1
    lines = [
        "# Product RefNr Missing Notes Fix",
        "",
        "This auxiliary file contains only Product final review rows still missing `final_human_notes`.",
        "",
        "## Inputs",
        "",
        f"- Final review workbook: `{workbook_path}`",
        f"- Validation report: `{validation_report}`",
        f"- Validation summary: `{validation_summary}`",
        "",
        "## Summary",
        "",
        f"- Missing notes included: {len(rows)}",
        "- Decisions are not applied automatically.",
        "- Protected files are not modified by this command.",
        "",
        "## Missing Notes By Decision",
        "",
    ]
    if by_decision:
        for decision, count in sorted(by_decision.items()):
            lines.append(f"- `{decision}`: {count}")
    else:
        lines.append("- No missing final notes found.")
    lines.extend(
        [
            "",
            "## Next Step",
            "",
            "Copy or adapt `final_human_notes_to_apply` into `product_refnr_final_review_required.xlsx`, then rerun `validate-product-refnr-final-review`.",
        ]
    )
    return "\n".join(lines)


def require_input(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"Required input file not found: {path}")


def run_product_refnr_missing_notes_fix(
    output_dir: Path = OUTPUT_DIR,
    workbook_path: Path | None = None,
    validation_report_path: Path | None = None,
    validation_summary_path: Path | None = None,
) -> ProductRefnrMissingNotesFixResult:
    current_run_id = run_id()
    workbook_path = workbook_path or output_dir / XLSX_NAME
    validation_report_path = validation_report_path or output_dir / REPORT_NAME
    validation_summary_path = validation_summary_path or output_dir / SUMMARY_NAME
    for path in (workbook_path, validation_report_path, validation_summary_path):
        require_input(path)

    workbook = load_workbook(workbook_path, data_only=True)
    rows = collect_missing_note_rows(workbook)

    xlsx_path = output_dir / XLSX_OUTPUT_NAME
    csv_path = output_dir / CSV_OUTPUT_NAME
    readme_path = output_dir / README_OUTPUT_NAME
    write_workbook(xlsx_path, rows, current_run_id)
    write_csv(csv_path, rows, current_run_id)
    backup_existing(readme_path, current_run_id)
    readme_path.write_text(render_readme(rows, workbook_path, validation_report_path, validation_summary_path), encoding="utf-8")

    return ProductRefnrMissingNotesFixResult(
        output_dir=output_dir,
        xlsx_path=xlsx_path,
        csv_path=csv_path,
        readme_path=readme_path,
        missing_notes_count=len(rows),
    )
