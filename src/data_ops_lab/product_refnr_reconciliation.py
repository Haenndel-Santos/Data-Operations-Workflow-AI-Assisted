from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .io_utils import normalize_columns, read_table
from .product_reference_audit import OUTPUT_DIR, PD_PATTERN, clean_value, product_dataframe
from .source_onboarding import backup_existing, ensure_dir, file_sha256


DB_DIR = Path("db")
DATA_DIR = Path("originaldatabase")
SUPPORTED_EXTENSIONS = {".xlsx", ".xls", ".csv"}
WORKBOOK_NAME = "product_refnr_reconciliation_review.xlsx"
REPORT_NAME = "product_refnr_reconciliation_report.md"
SCHEMA_REPORT_NAME = "product_refnr_schema_inspection.md"


@dataclass(frozen=True)
class ProductRefnrReconciliationResult:
    output_dir: Path
    source_path: Path
    workbook_path: Path
    report_path: Path
    schema_report_path: Path
    original_rows: int
    product_refnr_rows: int
    matched_rows: int
    corrected_refnr_rows: int
    conflict_rows: int
    unmatched_original_rows: int
    unmatched_refnr_rows: int
    product_finalized: bool


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def normalize_key(value: Any) -> str:
    return clean_value(value).casefold()


def source_row_number(index: Any) -> int:
    return int(index) + 2


def locate_product_refnr_file(db_dir: Path = DB_DIR, data_dir: Path = DATA_DIR) -> Path:
    search_dirs = [db_dir, data_dir]
    candidates: list[Path] = []
    for directory in search_dirs:
        if not directory.exists():
            continue
        for path in directory.iterdir():
            if not path.is_file() or path.suffix.lower() not in SUPPORTED_EXTENSIONS:
                continue
            stem = path.stem.casefold().replace(" ", "").replace("_", "").replace("-", "")
            name = path.name.casefold().replace(" ", "")
            if "product" in stem and ("ref.nr" in name or "refnr" in stem):
                candidates.append(path)
    if not candidates:
        raise FileNotFoundError("Could not find Product_ref.nr spreadsheet in db/ or originaldatabase/.")
    return sorted(candidates, key=lambda path: (0 if path.parent == db_dir else 1, path.name.casefold()))[0]


def sheet_names(path: Path) -> list[str]:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return pd.ExcelFile(path).sheet_names
    return ["csv"]


def read_product_refnr(path: Path) -> tuple[pd.DataFrame, str]:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        sheets = sheet_names(path)
        sheet = sheets[0]
        return normalize_columns(read_table(path, sheet_name=sheet)), sheet
    return normalize_columns(read_table(path)), "csv"


def detect_column(columns: list[str], candidates: tuple[str, ...], contains: tuple[str, ...] = ()) -> str | None:
    lowered = {column: column.lower() for column in columns}
    for candidate in candidates:
        if candidate in columns:
            return candidate
    for column, lowered_column in lowered.items():
        if contains and all(token in lowered_column for token in contains):
            return column
    return None


def detect_columns(df: pd.DataFrame) -> dict[str, str | None]:
    columns = list(df.columns)
    return {
        "product_ref_nr": detect_column(
            columns,
            ("product_ref_nr", "ref_nr", "pd_ref_nr", "corrected_ref_nr", "corrected_reference", "ref"),
            ("ref", "nr"),
        ),
        "part_nr_sku": detect_column(
            columns,
            ("part_nr_sku", "part_number_sku", "part_nr", "sku", "product_sku"),
            ("part", "sku"),
        ),
        "description": detect_column(columns, ("item_description", "description", "product_description"), ("description",)),
        "name": detect_column(columns, ("name", "product_name"), ("name",)),
        "product_group": detect_column(columns, ("product_group_name", "product_group", "category"), ("group",)),
        "supplier_ref": detect_column(columns, ("suppl_part_nr_sku", "supplier_part_nr", "supplier_sku"), ("suppl", "sku")),
        "supplier_name": detect_column(columns, ("creditor", "supplier", "supplier_name"), ("creditor",)),
    }


def duplicate_count(series: pd.Series) -> int:
    cleaned = series.map(clean_value)
    non_empty = cleaned[cleaned != ""]
    counts = non_empty.value_counts()
    return int((counts[counts > 1] - 1).sum())


def schema_inspection(path: Path, df: pd.DataFrame, selected_sheet: str) -> dict[str, Any]:
    empty_values = {column: int(df[column].map(clean_value).eq("").sum()) for column in df.columns}
    duplicate_values = {column: duplicate_count(df[column]) for column in df.columns}
    candidate_keys = []
    possible_join_columns = []
    for column in df.columns:
        cleaned = df[column].map(clean_value)
        non_empty = cleaned[cleaned != ""]
        non_empty_count = int(len(non_empty))
        unique_count = int(non_empty.nunique())
        if non_empty_count and unique_count == non_empty_count:
            candidate_keys.append(column)
        lowered = column.lower()
        if any(token in lowered for token in ("part", "sku", "ref", "description", "name", "barcode", "creditor", "supplier", "group")):
            possible_join_columns.append(column)
    return {
        "file_name": path.name,
        "file_path": str(path),
        "file_hash": file_sha256(path),
        "sheet_names": sheet_names(path),
        "selected_sheet": selected_sheet,
        "columns": list(df.columns),
        "row_count": int(len(df)),
        "empty_values": empty_values,
        "duplicate_values": duplicate_values,
        "candidate_key_columns": candidate_keys,
        "possible_join_columns": possible_join_columns,
        "detected_columns": detect_columns(df),
    }


def key_for_row(row: pd.Series, column: str | None) -> str:
    if not column:
        return ""
    return normalize_key(row.get(column, ""))


def build_index(df: pd.DataFrame, columns: list[str]) -> dict[str, list[int]]:
    index: dict[str, list[int]] = defaultdict(list)
    for row_index, row in df.iterrows():
        parts = [key_for_row(row, column) for column in columns]
        if all(parts):
            index["|".join(parts)].append(row_index)
    return index


def base_payload(prefix: str, row: pd.Series, columns: dict[str, str | None]) -> dict[str, Any]:
    payload = {f"{prefix}_source_row_number": source_row_number(row.name)}
    for semantic, column in columns.items():
        if column:
            payload[f"{prefix}_{semantic}"] = clean_value(row.get(column, ""))
    return payload


def reconcile(original_df: pd.DataFrame, refnr_df: pd.DataFrame) -> dict[str, list[dict[str, Any]]]:
    original_columns = detect_columns(original_df)
    refnr_columns = detect_columns(refnr_df)
    part_col_original = original_columns["part_nr_sku"]
    part_col_refnr = refnr_columns["part_nr_sku"]
    ref_col = refnr_columns["product_ref_nr"]
    description_original = original_columns["description"]
    description_refnr = refnr_columns["description"]
    group_original = original_columns["product_group"]
    group_refnr = refnr_columns["product_group"]
    name_original = original_columns["name"]
    name_refnr = refnr_columns["name"]

    part_index = build_index(refnr_df, [part_col_refnr] if part_col_refnr else [])
    desc_group_index = build_index(refnr_df, [description_refnr, group_refnr] if description_refnr and group_refnr else [])
    desc_index = build_index(refnr_df, [description_refnr] if description_refnr else [])
    name_index = build_index(refnr_df, [name_refnr] if name_refnr else [])

    matched = []
    corrections = []
    conflicts = []
    matched_refnr_indices: set[int] = set()

    for _, original_row in original_df.iterrows():
        match_method = ""
        candidates: list[int] = []
        part_key = key_for_row(original_row, part_col_original)
        if part_key and part_key in part_index:
            candidates = part_index[part_key]
            match_method = "part_nr_sku"
        if not candidates and description_original and group_original:
            key = "|".join([key_for_row(original_row, description_original), key_for_row(original_row, group_original)])
            if key in desc_group_index:
                candidates = desc_group_index[key]
                match_method = "description_product_group"
        if not candidates and description_original:
            key = key_for_row(original_row, description_original)
            if key in desc_index:
                candidates = desc_index[key]
                match_method = "description"
        if not candidates and name_original:
            key = key_for_row(original_row, name_original)
            if key in name_index:
                candidates = name_index[key]
                match_method = "name"

        original_payload = base_payload("original", original_row, original_columns)
        if len(candidates) == 1:
            ref_row = refnr_df.loc[candidates[0]]
            matched_refnr_indices.add(candidates[0])
            ref_payload = base_payload("refnr", ref_row, refnr_columns)
            corrected_ref = clean_value(ref_row.get(ref_col, "")) if ref_col else ""
            row = {
                **original_payload,
                **ref_payload,
                "match_method": match_method,
                "corrected_ref_nr": corrected_ref,
                "pd_ref_nr": corrected_ref if PD_PATTERN.match(corrected_ref) else "",
                "reference_status": "corrected_ref_nr_available" if corrected_ref else "missing_corrected_ref_nr",
                "conflict_reason": "",
            }
            matched.append(row)
            if corrected_ref:
                corrections.append(row)
            if ref_col and not corrected_ref:
                conflicts.append({**row, "conflict_reason": "missing_product_ref_nr_in_authoritative_file"})
        elif len(candidates) > 1:
            conflicts.append(
                {
                    **original_payload,
                    "match_method": match_method,
                    "candidate_count": len(candidates),
                    "conflict_reason": "multiple_product_refnr_rows_matched_original_product",
                }
            )

    matched_original_rows = {row["original_source_row_number"] for row in matched}
    conflict_original_rows = {row.get("original_source_row_number") for row in conflicts if row.get("original_source_row_number")}
    unmatched_original = [
        base_payload("original", row, original_columns)
        for _, row in original_df.iterrows()
        if source_row_number(row.name) not in matched_original_rows and source_row_number(row.name) not in conflict_original_rows
    ]
    unmatched_refnr = [
        base_payload("refnr", row, refnr_columns)
        for index, row in refnr_df.iterrows()
        if index not in matched_refnr_indices
    ]
    duplicates = duplicate_rows_in_refnr(refnr_df, refnr_columns)
    return {
        "matched_products": matched,
        "refnr_corrections": corrections,
        "conflicts": conflicts,
        "unmatched_original_product": unmatched_original,
        "unmatched_product_refnr": unmatched_refnr,
        "duplicates_in_product_refnr": duplicates,
    }


def duplicate_rows_in_refnr(df: pd.DataFrame, columns: dict[str, str | None]) -> list[dict[str, Any]]:
    rows = []
    for semantic in ("product_ref_nr", "part_nr_sku"):
        column = columns.get(semantic)
        if not column:
            continue
        cleaned = df[column].map(clean_value)
        counts = cleaned[cleaned != ""].value_counts()
        duplicate_values = counts[counts > 1]
        for value, count in duplicate_values.items():
            for _, row in df[cleaned == value].iterrows():
                rows.append(
                    {
                        "duplicate_column": column,
                        "duplicate_type": semantic,
                        "duplicate_value": value,
                        "duplicate_count": int(count),
                        **base_payload("refnr", row, columns),
                    }
                )
    return rows


def summary_rows(schema: dict[str, Any], reconciliation: dict[str, list[dict[str, Any]]], original_rows: int) -> list[dict[str, Any]]:
    return [
        {"metric": "product_refnr_file", "value": schema["file_name"]},
        {"metric": "product_refnr_rows", "value": schema["row_count"]},
        {"metric": "original_product_rows", "value": original_rows},
        {"metric": "matched_products", "value": len(reconciliation["matched_products"])},
        {"metric": "corrected_ref_nr_rows", "value": len(reconciliation["refnr_corrections"])},
        {"metric": "conflicts", "value": len(reconciliation["conflicts"])},
        {"metric": "unmatched_original_product", "value": len(reconciliation["unmatched_original_product"])},
        {"metric": "unmatched_product_refnr", "value": len(reconciliation["unmatched_product_refnr"])},
        {"metric": "duplicates_in_product_refnr", "value": len(reconciliation["duplicates_in_product_refnr"])},
        {"metric": "product_final_key_decision", "value": "not_finalized_pending_reconciliation_review"},
    ]


def write_sheet(sheet, rows: list[dict[str, Any]], fallback_columns: list[str]) -> None:
    columns = list(rows[0].keys()) if rows else fallback_columns
    sheet.append(columns)
    for row in rows:
        sheet.append([row.get(column, "") for column in columns])
    style_sheet(sheet)


def style_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.freeze_panes = "A2"
    if sheet.max_row >= 1 and sheet.max_column >= 1:
        sheet.auto_filter.ref = sheet.dimensions
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for column_cells in sheet.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)


def write_workbook(path: Path, schema: dict[str, Any], reconciliation: dict[str, list[dict[str, Any]]], original_rows: int) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    fallback = ["review_status", "human_decision", "human_notes"]
    for sheet_name in [
        "matched_products",
        "refnr_corrections",
        "conflicts",
        "unmatched_original_product",
        "unmatched_product_refnr",
        "duplicates_in_product_refnr",
    ]:
        sheet = workbook.create_sheet(sheet_name)
        write_sheet(sheet, reconciliation[sheet_name], fallback)
    summary = workbook.create_sheet("reconciliation_summary")
    write_sheet(summary, summary_rows(schema, reconciliation, original_rows), ["metric", "value"])
    decision_log = workbook.create_sheet("decision_log")
    write_sheet(
        decision_log,
        [],
        ["decision_date", "reviewer", "topic", "decision", "rationale", "follow_up_action"],
    )
    workbook.save(path)


def render_schema_report(schema: dict[str, Any]) -> str:
    lines = [
        "# Product_ref.nr Schema Inspection",
        "",
        f"- File name: `{schema['file_name']}`",
        f"- File path: `{schema['file_path']}`",
        f"- File SHA256: `{schema['file_hash']}`",
        f"- Sheet names: {', '.join(f'`{sheet}`' for sheet in schema['sheet_names'])}",
        f"- Selected sheet: `{schema['selected_sheet']}`",
        f"- Row count: {schema['row_count']}",
        "",
        "## Detected Columns",
        "",
    ]
    for column in schema["columns"]:
        lines.append(f"- `{column}`")
    lines.extend(["", "## Empty Values", ""])
    for column, count in schema["empty_values"].items():
        lines.append(f"- `{column}`: {count}")
    lines.extend(["", "## Duplicate Values", ""])
    for column, count in schema["duplicate_values"].items():
        lines.append(f"- `{column}`: {count}")
    lines.extend(["", "## Candidate Key Columns", ""])
    if schema["candidate_key_columns"]:
        for column in schema["candidate_key_columns"]:
            lines.append(f"- `{column}`")
    else:
        lines.append("- No fully unique non-empty candidate key columns detected.")
    lines.extend(["", "## Possible Join Columns", ""])
    for column in schema["possible_join_columns"]:
        lines.append(f"- `{column}`")
    lines.extend(["", "## Auto-Detected Semantic Columns", ""])
    for semantic, column in schema["detected_columns"].items():
        lines.append(f"- `{semantic}`: `{column or 'not_detected'}`")
    return "\n".join(lines)


def render_reconciliation_report(schema: dict[str, Any], reconciliation: dict[str, list[dict[str, Any]]], original_rows: int) -> str:
    matched = len(reconciliation["matched_products"])
    corrections = len(reconciliation["refnr_corrections"])
    conflicts = len(reconciliation["conflicts"])
    unmatched_original = len(reconciliation["unmatched_original_product"])
    unmatched_refnr = len(reconciliation["unmatched_product_refnr"])
    duplicate_refnr_rows = len(reconciliation["duplicates_in_product_refnr"])
    reliable_ref = bool(
        corrections
        and conflicts == 0
        and unmatched_original == 0
        and duplicate_refnr_rows == 0
        and schema["detected_columns"].get("product_ref_nr")
    )
    lines = [
        "# Product Ref.nr Reconciliation Report",
        "",
        "This step reads `Product_ref.nr` as an authoritative correction/enrichment source for Product references. It does not finalize Product key approval.",
        "",
        "## Source",
        "",
        f"- Product_ref.nr file: `{schema['file_path']}`",
        f"- Product_ref.nr rows: {schema['row_count']}",
        f"- Original Product rows: {original_rows}",
        "",
        "## Reconciliation Summary",
        "",
        f"- Matched Product rows: {matched}",
        f"- Rows receiving corrected `ref_nr`: {corrections}",
        f"- Conflicts remaining: {conflicts}",
        f"- Unmatched original Product rows: {unmatched_original}",
        f"- Unmatched Product_ref.nr rows: {unmatched_refnr}",
        f"- Duplicate rows in Product_ref.nr review set: {duplicate_refnr_rows}",
        "",
        "## Previous Product Issues",
        "",
        "- Duplicate, empty, and non-PD Product reference issues are not finalized by this step.",
        "- `Product_ref.nr` provides a candidate corrected `product_ref_nr` that must be reviewed before the final Product key decision is regenerated.",
        f"- Does Product_ref.nr currently resolve all prior Product reference issues automatically? {'Yes' if reliable_ref else 'No, reconciliation review is still required.'}",
        "",
        "## Modeling Recommendation Pending Review",
        "",
        "- `product_id`: technical/generated primary key",
        "- `product_ref_nr`: corrected canonical product reference from `Product_ref.nr`",
        "- `part_nr_sku`: business/search/customer/supplier-facing product reference",
        "- `pd_ref_nr`: optional PD-style serial reference where applicable",
        "- `source_record_id`: original source row identifier",
        "- `review_status`: validation status",
        "",
        "## Product Key Decision Status",
        "",
        "- Product final key decision: `not_finalized_pending_reconciliation_review`",
        "- Do not update `approved_keys.yml` or `approved_relationships.yml` from this reconciliation.",
    ]
    return "\n".join(lines)


def run_product_refnr_reconciliation(
    db_dir: Path = DB_DIR,
    data_dir: Path = DATA_DIR,
    output_dir: Path = OUTPUT_DIR,
) -> ProductRefnrReconciliationResult:
    current_run_id = run_id()
    source_path = locate_product_refnr_file(db_dir, data_dir)
    refnr_df, selected_sheet = read_product_refnr(source_path)
    original_df = product_dataframe(data_dir)
    schema = schema_inspection(source_path, refnr_df, selected_sheet)
    reconciliation = reconcile(original_df, refnr_df)

    ensure_dir(output_dir)
    workbook_path = output_dir / WORKBOOK_NAME
    report_path = output_dir / REPORT_NAME
    schema_report_path = output_dir / SCHEMA_REPORT_NAME
    backup_existing(workbook_path, current_run_id)
    backup_existing(report_path, current_run_id)
    backup_existing(schema_report_path, current_run_id)
    write_workbook(workbook_path, schema, reconciliation, len(original_df))
    report_path.write_text(render_reconciliation_report(schema, reconciliation, len(original_df)), encoding="utf-8")
    schema_report_path.write_text(render_schema_report(schema), encoding="utf-8")

    return ProductRefnrReconciliationResult(
        output_dir=output_dir,
        source_path=source_path,
        workbook_path=workbook_path,
        report_path=report_path,
        schema_report_path=schema_report_path,
        original_rows=len(original_df),
        product_refnr_rows=len(refnr_df),
        matched_rows=len(reconciliation["matched_products"]),
        corrected_refnr_rows=len(reconciliation["refnr_corrections"]),
        conflict_rows=len(reconciliation["conflicts"]),
        unmatched_original_rows=len(reconciliation["unmatched_original_product"]),
        unmatched_refnr_rows=len(reconciliation["unmatched_product_refnr"]),
        product_finalized=False,
    )
