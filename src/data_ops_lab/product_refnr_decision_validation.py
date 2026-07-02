from __future__ import annotations

import csv
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .product_reference_audit import OUTPUT_DIR, clean_value
from .product_refnr_human_review import DECISION_OPTIONS, SHORTLIST_WORKBOOK
from .source_onboarding import backup_existing, ensure_dir


REPORT_NAME = "product_refnr_human_decision_validation_report.md"
SUMMARY_NAME = "product_refnr_human_decision_summary.csv"
REVIEW_SHEETS = ["Conflicts", "Unmatched Original Product", "Unmatched Product RefNr", "Duplicate RefNr Review"]
ALLOWED_DECISIONS = set(DECISION_OPTIONS)
PENDING_DECISION = "pending"
APPROVED_DECISIONS = {
    "approved_use_corrected_product_ref_nr",
    "approved_keep_original_part_nr_sku_only",
    "approved_create_technical_product_id_only",
    "merge_duplicate_records",
    "keep_as_separate_products",
}


@dataclass(frozen=True)
class ProductRefnrDecisionValidationResult:
    output_dir: Path
    report_path: Path
    summary_csv_path: Path
    total_decisions: int
    valid_decisions: int
    pending_decisions: int
    invalid_decisions: int
    missing_notes: int
    recommended_next_step: str


def run_id() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def rows_from_sheet(workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        return []
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [clean_value(value) for value in rows[0]]
    output = []
    for raw_row in rows[1:]:
        if not any(clean_value(value) for value in raw_row):
            continue
        output.append({header: clean_value(value) for header, value in zip(headers, raw_row, strict=False) if header})
    return output


def normalized_decision(row: dict[str, Any]) -> str:
    return clean_value(row.get("human_decision", "")).casefold()


def note(row: dict[str, Any]) -> str:
    return clean_value(row.get("human_notes", ""))


def issue_id(row: dict[str, Any], fallback: int) -> str:
    return clean_value(row.get("issue_id", "")) or f"row_{fallback}"


def requires_note(row: dict[str, Any]) -> bool:
    issue_type = clean_value(row.get("issue_type", ""))
    decision = normalized_decision(row)
    if not decision or decision == PENDING_DECISION or decision not in ALLOWED_DECISIONS:
        return False
    if issue_type == "conflict" and decision.startswith("approved_"):
        return True
    if issue_type == "duplicate_refnr_review" and decision in {
        "approved_use_corrected_product_ref_nr",
        "merge_duplicate_records",
        "keep_as_separate_products",
    }:
        return True
    if issue_type == "unmatched_product_refnr" and decision == "approved_use_corrected_product_ref_nr":
        return True
    return False


def inconsistency(row: dict[str, Any]) -> str:
    issue_type = clean_value(row.get("issue_type", ""))
    recommended = clean_value(row.get("recommended_action", "")).casefold()
    decision = normalized_decision(row)
    if decision not in ALLOWED_DECISIONS or decision in {"", PENDING_DECISION}:
        return ""
    if issue_type == "conflict" and recommended == "needs_business_context" and decision == "approved_keep_original_part_nr_sku_only":
        return "conflict_keeps_original_without_corrected_mapping"
    if issue_type == "unmatched_original_product" and decision == "approved_use_corrected_product_ref_nr":
        return "unmatched_original_has_no_corrected_refnr_to_use"
    if issue_type == "unmatched_product_refnr" and decision == "approved_keep_original_part_nr_sku_only":
        return "unmatched_refnr_has_no_original_product_to_keep"
    if issue_type == "duplicate_refnr_review" and decision == "approved_use_corrected_product_ref_nr" and not note(row):
        return "duplicate_refnr_approved_without_duplicate_resolution_note"
    return ""


def validate_rows(rows_by_sheet: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    summary_by_issue_type: dict[str, Counter[str]] = defaultdict(Counter)
    invalid_entries = []
    empty_entries = []
    pending_entries = []
    missing_note_entries = []
    inconsistency_entries = []
    total = 0
    valid = 0

    for sheet_name, rows in rows_by_sheet.items():
        for index, row in enumerate(rows, start=2):
            total += 1
            issue_type = clean_value(row.get("issue_type", "")) or sheet_name
            decision = normalized_decision(row)
            row_id = issue_id(row, index)
            counter = summary_by_issue_type[issue_type]
            counter["total_rows"] += 1
            if decision in ALLOWED_DECISIONS:
                valid += 1
                if decision == PENDING_DECISION:
                    counter["pending_count"] += 1
                    pending_entries.append({"sheet": sheet_name, "issue_id": row_id, "decision": decision})
                elif decision == "needs_business_context":
                    counter["needs_context_count"] += 1
                elif decision == "rejected":
                    counter["rejected_count"] += 1
                else:
                    counter["approved_count"] += 1
            elif not decision:
                counter["invalid_decision_count"] += 1
                empty_entries.append({"sheet": sheet_name, "issue_id": row_id, "decision": ""})
            else:
                counter["invalid_decision_count"] += 1
                invalid_entries.append({"sheet": sheet_name, "issue_id": row_id, "decision": decision})

            if requires_note(row) and not note(row):
                counter["missing_notes_count"] += 1
                missing_note_entries.append({"sheet": sheet_name, "issue_id": row_id, "decision": decision})
            problem = inconsistency(row)
            if problem:
                inconsistency_entries.append(
                    {
                        "sheet": sheet_name,
                        "issue_id": row_id,
                        "decision": decision,
                        "problem": problem,
                    }
                )

    return {
        "summary_by_issue_type": summary_by_issue_type,
        "invalid_entries": invalid_entries,
        "empty_entries": empty_entries,
        "pending_entries": pending_entries,
        "missing_note_entries": missing_note_entries,
        "inconsistency_entries": inconsistency_entries,
        "total_decisions": total,
        "valid_decisions": valid,
        "pending_decisions": len(pending_entries),
        "invalid_decisions": len(invalid_entries) + len(empty_entries),
        "missing_notes": len(missing_note_entries),
    }


def write_summary_csv(path: Path, summary_by_issue_type: dict[str, Counter[str]], current_run_id: str) -> None:
    ensure_dir(path.parent)
    backup_existing(path, current_run_id)
    columns = [
        "issue_type",
        "total_rows",
        "approved_count",
        "needs_context_count",
        "rejected_count",
        "pending_count",
        "invalid_decision_count",
        "missing_notes_count",
    ]
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for issue_type in sorted(summary_by_issue_type):
            counts = summary_by_issue_type[issue_type]
            writer.writerow({column: issue_type if column == "issue_type" else int(counts.get(column, 0)) for column in columns})


def render_entries(title: str, entries: list[dict[str, str]], empty_text: str) -> list[str]:
    lines = ["", f"## {title}", ""]
    if not entries:
        lines.append(f"- {empty_text}")
    else:
        for entry in entries[:50]:
            detail = f"; problem={entry['problem']}" if "problem" in entry else ""
            lines.append(f"- {entry['sheet']} / {entry['issue_id']}: decision=`{entry['decision']}`{detail}")
        if len(entries) > 50:
            lines.append(f"- ... {len(entries) - 50} more rows not shown")
    return lines


def next_step(validation: dict[str, Any]) -> str:
    if validation["invalid_decisions"] or validation["pending_decisions"] or validation["missing_notes"] or validation["inconsistency_entries"]:
        return "Do not apply final decisions yet. Resolve invalid/pending decisions, required notes, and inconsistencies first."
    return "Step 3E.4 — Apply Product Reconciliation Decisions"


def render_report(validation: dict[str, Any]) -> str:
    lines = [
        "# Product RefNr Human Decision Validation Report",
        "",
        "This validation only reads the completed Product RefNr shortlist workbook. No Product decision is applied.",
        "",
        "## Summary",
        "",
        f"- Total decisions read: {validation['total_decisions']}",
        f"- Valid decisions: {validation['valid_decisions']}",
        f"- Pending decisions: {validation['pending_decisions']}",
        f"- Invalid or empty decisions: {validation['invalid_decisions']}",
        f"- Missing required notes: {validation['missing_notes']}",
        f"- Inconsistencies: {len(validation['inconsistency_entries'])}",
        f"- Recommended next step: {next_step(validation)}",
        "",
        "## Decision Counts By Issue Type",
        "",
    ]
    for issue_type, counts in sorted(validation["summary_by_issue_type"].items()):
        lines.append(
            "- {issue_type}: total={total}, approved={approved}, needs_context={context}, rejected={rejected}, pending={pending}, invalid={invalid}, missing_notes={missing}".format(
                issue_type=issue_type,
                total=int(counts.get("total_rows", 0)),
                approved=int(counts.get("approved_count", 0)),
                context=int(counts.get("needs_context_count", 0)),
                rejected=int(counts.get("rejected_count", 0)),
                pending=int(counts.get("pending_count", 0)),
                invalid=int(counts.get("invalid_decision_count", 0)),
                missing=int(counts.get("missing_notes_count", 0)),
            )
        )
    lines.extend(render_entries("Invalid Decisions", validation["invalid_entries"], "No invalid decision values."))
    lines.extend(render_entries("Empty Decisions", validation["empty_entries"], "No empty decisions."))
    lines.extend(render_entries("Pending Decisions", validation["pending_entries"], "No pending decisions."))
    lines.extend(render_entries("Missing Required Notes", validation["missing_note_entries"], "No required notes are missing."))
    lines.extend(render_entries("Decision Inconsistencies", validation["inconsistency_entries"], "No decision inconsistencies detected."))
    lines.extend(
        [
            "",
            "## Guardrails",
            "",
            "- `approved_keys.yml` was not updated.",
            "- `approved_relationships.yml` was not updated.",
            "- Product final key decision was not applied.",
        ]
    )
    return "\n".join(lines)


def run_validate_product_refnr_decisions(
    output_dir: Path = OUTPUT_DIR,
    workbook_path: Path | None = None,
) -> ProductRefnrDecisionValidationResult:
    current_run_id = run_id()
    workbook_path = workbook_path or output_dir / SHORTLIST_WORKBOOK
    if not workbook_path.exists():
        raise FileNotFoundError(f"Product RefNr human review shortlist not found: {workbook_path}")
    workbook = load_workbook(workbook_path, data_only=True)
    rows_by_sheet = {sheet: rows_from_sheet(workbook, sheet) for sheet in REVIEW_SHEETS}
    validation = validate_rows(rows_by_sheet)

    report_path = output_dir / REPORT_NAME
    summary_csv_path = output_dir / SUMMARY_NAME
    ensure_dir(output_dir)
    backup_existing(report_path, current_run_id)
    report_path.write_text(render_report(validation), encoding="utf-8")
    write_summary_csv(summary_csv_path, validation["summary_by_issue_type"], current_run_id)

    return ProductRefnrDecisionValidationResult(
        output_dir=output_dir,
        report_path=report_path,
        summary_csv_path=summary_csv_path,
        total_decisions=validation["total_decisions"],
        valid_decisions=validation["valid_decisions"],
        pending_decisions=validation["pending_decisions"],
        invalid_decisions=validation["invalid_decisions"],
        missing_notes=validation["missing_notes"],
        recommended_next_step=next_step(validation),
    )
