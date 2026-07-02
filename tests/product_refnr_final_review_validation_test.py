from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

from data_ops_lab.product_refnr_final_review_spreadsheet import REVIEW_COLUMNS
from data_ops_lab.product_refnr_final_review_validation import run_validate_product_refnr_final_review


def add_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(REVIEW_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in REVIEW_COLUMNS])


def review_row(
    review_id: str,
    issue_type: str,
    decision: str,
    notes: str,
    recommended_action: str = "needs_business_context",
    problem_type: str = "missing_note",
) -> dict[str, str]:
    return {
        "review_id": review_id,
        "original_sheet": "Conflicts",
        "original_excel_row": "2",
        "issue_id": review_id.replace("REVIEW", "ISSUE"),
        "issue_type": issue_type,
        "product_original_identifier": "original_row_1",
        "product_refnr_identifier": "refnr_row_1",
        "original_part_nr_sku": "SKU1",
        "corrected_product_ref_nr": "PD2600001",
        "optional_pd_ref_nr": "PD2600001",
        "current_human_decision": decision,
        "current_human_notes": notes,
        "problem_type": problem_type,
        "problem_explanation": "Requires final note.",
        "suggested_human_decision": decision,
        "suggested_human_notes": "",
        "required_action": "Fill final_human_notes.",
        "final_human_decision": decision,
        "final_human_notes": notes,
    }


def build_clean_final_review(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("README").append(["readme"])
    required_row = review_row("REVIEW_001", "conflict", "approved_use_corrected_product_ref_nr", "")
    missing_row = review_row(
        "REVIEW_001",
        "conflict",
        "approved_use_corrected_product_ref_nr",
        "Corrected reference approved after final manual review.",
    )
    duplicate_row = review_row(
        "REVIEW_002",
        "duplicate_refnr_review",
        "merge_duplicate_records",
        "Duplicate references confirmed as same product and should be merged.",
    )
    add_sheet(workbook, "Required Review", [required_row, duplicate_row])
    add_sheet(workbook, "Missing Notes", [missing_row])
    add_sheet(workbook, "Inconsistencies", [])
    add_sheet(workbook, "All Product Exceptions", [missing_row, duplicate_row])
    workbook.save(path)


def build_blocked_final_review(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("README").append(["readme"])
    invalid = review_row("REVIEW_003", "conflict", "not_allowed", "bad")
    missing_note = review_row("REVIEW_004", "conflict", "approved_use_corrected_product_ref_nr", "")
    inconsistent = review_row(
        "REVIEW_005",
        "conflict",
        "approved_keep_original_part_nr_sku_only",
        "Kept original despite conflict.",
        problem_type="inconsistency",
    )
    add_sheet(workbook, "Required Review", [invalid, missing_note, inconsistent])
    add_sheet(workbook, "Missing Notes", [])
    add_sheet(workbook, "Inconsistencies", [inconsistent])
    add_sheet(workbook, "All Product Exceptions", [invalid, missing_note, inconsistent])
    workbook.save(path)


def test_validate_product_refnr_final_review_clean_generates_validated_workbook_and_preserves_files(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    workbook_path = output_dir / "product_refnr_final_review_required.xlsx"
    build_clean_final_review(workbook_path)

    product = data_dir / "Product.xlsx"
    product_refnr = data_dir / "Product_ref.nr.xlsx"
    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    product.write_bytes(b"product")
    product_refnr.write_bytes(b"refnr")
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    protected_before = {
        product: product.read_bytes(),
        product_refnr: product_refnr.read_bytes(),
        approved_keys: approved_keys.read_text(encoding="utf-8"),
        approved_relationships: approved_relationships.read_text(encoding="utf-8"),
    }

    result = run_validate_product_refnr_final_review(output_dir, workbook_path)

    assert result.ready_for_apply is True
    assert result.total_decisions == 2
    assert result.valid_decisions == 2
    assert result.missing_notes == 0
    assert result.inconsistencies == 0
    assert result.validated_workbook_path is not None
    assert result.validated_workbook_path.exists()
    assert result.report_path.exists()
    assert result.summary_csv_path.exists()
    assert product.read_bytes() == protected_before[product]
    assert product_refnr.read_bytes() == protected_before[product_refnr]
    assert approved_keys.read_text(encoding="utf-8") == protected_before[approved_keys]
    assert approved_relationships.read_text(encoding="utf-8") == protected_before[approved_relationships]

    validated = load_workbook(result.validated_workbook_path, data_only=True)
    assert "Required Review" in validated.sheetnames
    report = result.report_path.read_text(encoding="utf-8")
    assert "Ready for apply: true" in report


def test_validate_product_refnr_final_review_detects_blockers(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    output_dir.mkdir()
    workbook_path = output_dir / "product_refnr_final_review_required.xlsx"
    build_blocked_final_review(workbook_path)

    result = run_validate_product_refnr_final_review(output_dir, workbook_path)

    assert result.ready_for_apply is False
    assert result.total_decisions == 3
    assert result.invalid_decisions == 1
    assert result.missing_notes == 1
    assert result.inconsistencies == 1
    assert result.validated_workbook_path is None

    with result.summary_csv_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    conflict = next(row for row in rows if row["issue_type"] == "conflict")
    assert conflict["invalid_decision_count"] == "1"
    assert conflict["missing_notes_count"] == "1"
    assert conflict["inconsistency_count"] == "1"
