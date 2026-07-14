from __future__ import annotations

import csv
from pathlib import Path

import pandas as pd
import pytest
import yaml
from openpyxl import Workbook

from data_ops_lab.product_materialization import run_product_materialization
from data_ops_lab.product_refnr_application import run_product_refnr_application
from data_ops_lab.product_refnr_final_review_spreadsheet import REVIEW_COLUMNS


def review_row(
    review_id: str,
    issue_id: str,
    issue_type: str,
    decision: str,
    *,
    original_identifier: str = "",
    refnr_identifier: str = "",
) -> dict[str, str]:
    return {
        "review_id": review_id,
        "original_sheet": issue_type,
        "original_excel_row": "2",
        "issue_id": issue_id,
        "issue_type": issue_type,
        "product_original_identifier": original_identifier,
        "product_refnr_identifier": refnr_identifier,
        "original_part_nr_sku": "fixture-value",
        "corrected_product_ref_nr": "fixture-corrected-value",
        "optional_pd_ref_nr": "",
        "current_human_decision": decision,
        "current_human_notes": "Approved fixture decision.",
        "problem_type": "",
        "problem_explanation": "",
        "suggested_human_decision": decision,
        "suggested_human_notes": "",
        "required_action": "",
        "final_human_decision": decision,
        "final_human_notes": "Approved fixture decision.",
    }


def write_review_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("README").append(["readme"])
    for title in ("Required Review", "Missing Notes", "Inconsistencies", "All Product Exceptions"):
        sheet = workbook.create_sheet(title)
        sheet.append(REVIEW_COLUMNS)
        if title == "All Product Exceptions":
            for row in rows:
                sheet.append([row.get(column, "") for column in REVIEW_COLUMNS])
    workbook.save(path)


def write_product_sources(data_dir: Path) -> tuple[Path, Path]:
    product = pd.DataFrame(
        [
            {"Part nr. (SKU)": "SKU-A", "Item description": "Alpha", "Name": "Alpha"},
            {"Part nr. (SKU)": "SKU-DUP", "Item description": "Beta", "Name": "Beta"},
            {"Part nr. (SKU)": "SKU-REJECT", "Item description": "Rejected", "Name": "Rejected"},
        ]
    )
    refnr = pd.DataFrame(
        [
            {"Ref. nr.": "PD2600001", "Part nr. (SKU)": "SKU-A", "Item description": "Alpha", "Name": "Alpha"},
            {"Ref. nr.": "PD2600002", "Part nr. (SKU)": "SKU-DUP", "Item description": "Beta", "Name": "Beta"},
            {"Ref. nr.": "PD2600003", "Part nr. (SKU)": "SKU-DUP", "Item description": "Other", "Name": "Other"},
            {"Ref. nr.": "PD2600004", "Part nr. (SKU)": "SKU-NEW", "Item description": "New", "Name": "New"},
        ]
    )
    product_path = data_dir / "Product.xlsx"
    refnr_path = data_dir / "Product_ref.nr.xlsx"
    with pd.ExcelWriter(product_path, engine="openpyxl") as writer:
        product.to_excel(writer, sheet_name="Export_Product", index=False)
    with pd.ExcelWriter(refnr_path, engine="openpyxl") as writer:
        refnr.to_excel(writer, sheet_name="Export_Product", index=False)
    return product_path, refnr_path


def applied_fixture(tmp_path: Path) -> tuple[Path, Path, Path, Path, Path, Path]:
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    application_output = tmp_path / "application"
    materialization_output = tmp_path / "materialization"
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    application_output.mkdir()
    product_path, refnr_path = write_product_sources(data_dir)

    rows = [
        review_row(
            "REVIEW_001",
            "CONFLICT_001",
            "conflict",
            "approved_use_corrected_product_ref_nr",
            original_identifier="original_row_3",
        ),
        review_row(
            "REVIEW_002",
            "UNMATCHED_ORIGINAL_001",
            "unmatched_original_product",
            "rejected",
            original_identifier="original_row_4",
        ),
        review_row(
            "REVIEW_003",
            "UNMATCHED_REFNR_001",
            "unmatched_product_refnr",
            "approved_use_corrected_product_ref_nr",
            refnr_identifier="refnr_row_3",
        ),
        review_row(
            "REVIEW_004",
            "UNMATCHED_REFNR_002",
            "unmatched_product_refnr",
            "approved_use_corrected_product_ref_nr",
            refnr_identifier="refnr_row_4",
        ),
        review_row(
            "REVIEW_005",
            "UNMATCHED_REFNR_003",
            "unmatched_product_refnr",
            "approved_use_corrected_product_ref_nr",
            refnr_identifier="refnr_row_5",
        ),
        review_row(
            "REVIEW_006",
            "DUPLICATE_REFNR_001",
            "duplicate_refnr_review",
            "approved_use_corrected_product_ref_nr",
            refnr_identifier="refnr_row_3",
        ),
        review_row(
            "REVIEW_007",
            "DUPLICATE_REFNR_002",
            "duplicate_refnr_review",
            "rejected",
            refnr_identifier="refnr_row_4",
        ),
    ]
    workbook_path = tmp_path / "validated.xlsx"
    write_review_workbook(workbook_path, rows)
    application = run_product_refnr_application(
        workbook_path,
        application_output,
        config_dir,
        apply=True,
    )
    return data_dir, workbook_path, application.state_path, materialization_output, product_path, refnr_path


def test_product_materialization_builds_preview_and_is_idempotent(tmp_path: Path) -> None:
    data_dir, workbook_path, state_path, output_dir, product_path, refnr_path = applied_fixture(tmp_path)
    protected = {
        path: path.read_bytes()
        for path in (product_path, refnr_path, workbook_path, state_path)
    }

    first = run_product_materialization(data_dir, workbook_path, state_path, output_dir)
    first_files = {path.name: path.read_bytes() for path in output_dir.iterdir() if path.is_file()}
    second = run_product_materialization(data_dir, workbook_path, state_path, output_dir)

    assert first.status == "ready_for_local_preview"
    assert first.target_rows == 3
    assert first.excluded_identifiers == 2
    assert first.blocker_count == 0
    assert first.outputs_changed is True
    assert second.outputs_changed is False
    assert first.preview_path is not None
    assert first.lineage_path is not None
    assert first.exclusions_path is not None
    assert first_files == {path.name: path.read_bytes() for path in output_dir.iterdir() if path.is_file()}
    assert all(path.read_bytes() == content for path, content in protected.items())

    with first.preview_path.open(newline="", encoding="utf-8") as handle:
        preview = list(csv.DictReader(handle))
    with first.lineage_path.open(newline="", encoding="utf-8") as handle:
        lineage = list(csv.DictReader(handle))
    with first.exclusions_path.open(newline="", encoding="utf-8") as handle:
        exclusions = list(csv.DictReader(handle))
    assert len(preview) == len(lineage) == 3
    assert len({row["product_id"] for row in preview}) == 3
    assert all(row["product_ref_nr"] for row in preview)
    assert {row["materialization_action"] for row in lineage} == {
        "matched_authoritative_correction",
        "approved_same_row_conflict_resolution",
        "approved_product_refnr_only",
    }
    assert {row["source_identifier"] for row in exclusions} == {
        "original_row_4",
        "refnr_row_4",
    }


def test_product_materialization_blocks_approved_empty_authoritative_row(tmp_path: Path) -> None:
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    application_output = tmp_path / "application"
    output_dir = tmp_path / "materialization"
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    application_output.mkdir()
    product = pd.DataFrame(
        [
            {"Part nr. (SKU)": "SKU-A", "Item description": "Alpha"},
            {"Part nr. (SKU)": None, "Item description": None},
            {"Part nr. (SKU)": "SKU-C", "Item description": "Charlie"},
        ]
    )
    refnr = pd.DataFrame(
        [
            {"Ref. nr.": "PD2600001", "Part nr. (SKU)": "SKU-A", "Item description": "Alpha"},
            {"Ref. nr.": None, "Part nr. (SKU)": None, "Item description": None},
            {"Ref. nr.": "PD2600003", "Part nr. (SKU)": "SKU-C", "Item description": "Charlie"},
        ]
    )
    with pd.ExcelWriter(data_dir / "Product.xlsx", engine="openpyxl") as writer:
        product.to_excel(writer, sheet_name="Export_Product", index=False)
    with pd.ExcelWriter(data_dir / "Product_ref.nr.xlsx", engine="openpyxl") as writer:
        refnr.to_excel(writer, sheet_name="Export_Product", index=False)
    workbook_path = tmp_path / "validated.xlsx"
    write_review_workbook(
        workbook_path,
        [
            review_row(
                "REVIEW_001",
                "UNMATCHED_ORIGINAL_001",
                "unmatched_original_product",
                "rejected",
                original_identifier="original_row_3",
            ),
            review_row(
                "REVIEW_002",
                "UNMATCHED_REFNR_001",
                "unmatched_product_refnr",
                "approved_use_corrected_product_ref_nr",
                refnr_identifier="refnr_row_3",
            ),
        ],
    )
    application = run_product_refnr_application(
        workbook_path,
        application_output,
        config_dir,
        apply=True,
    )

    first = run_product_materialization(data_dir, workbook_path, application.state_path, output_dir)
    second = run_product_materialization(data_dir, workbook_path, application.state_path, output_dir)

    assert first.status == "blocked"
    assert first.blocker_count == 1
    assert first.target_rows == 0
    assert first.preview_path is None
    assert first.outputs_changed is True
    assert second.outputs_changed is False
    with first.blockers_path.open(newline="", encoding="utf-8") as handle:
        blockers = list(csv.DictReader(handle))
    assert blockers[0]["blocker_type"] == "approved_authoritative_row_empty"
    assert blockers[0]["source_identifier"] == "refnr_row_3"
    manifest = yaml.safe_load(first.manifest_path.read_text(encoding="utf-8"))
    assert manifest["status"] == "blocked"
    assert manifest["counts"]["blockers"] == 1


def test_product_materialization_refuses_state_workbook_mismatch(tmp_path: Path) -> None:
    data_dir, workbook_path, state_path, output_dir, _, _ = applied_fixture(tmp_path)
    changed_workbook = tmp_path / "changed.xlsx"
    write_review_workbook(
        changed_workbook,
        [
            review_row(
                "REVIEW_001",
                "CHANGED_001",
                "conflict",
                "rejected",
                original_identifier="original_row_3",
            )
        ],
    )

    with pytest.raises(ValueError, match="does not exactly match"):
        run_product_materialization(data_dir, changed_workbook, state_path, output_dir)

    assert not output_dir.exists()


def test_product_materialization_refuses_different_existing_outputs(tmp_path: Path) -> None:
    data_dir, workbook_path, state_path, output_dir, _, _ = applied_fixture(tmp_path)
    first = run_product_materialization(data_dir, workbook_path, state_path, output_dir)
    first.manifest_path.write_text("changed generated evidence\n", encoding="utf-8")

    with pytest.raises(ValueError, match="existing generated evidence was not overwritten"):
        run_product_materialization(data_dir, workbook_path, state_path, output_dir)

    assert first.manifest_path.read_text(encoding="utf-8") == "changed generated evidence\n"


def test_product_materialization_blocks_source_identifier_out_of_range(tmp_path: Path) -> None:
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    application_output = tmp_path / "application"
    output_dir = tmp_path / "materialization"
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    application_output.mkdir()
    product = pd.DataFrame([{"Part nr. (SKU)": "SKU-A", "Item description": "Alpha"}])
    refnr = pd.DataFrame(
        [{"Ref. nr.": "PD2600001", "Part nr. (SKU)": "SKU-A", "Item description": "Alpha"}]
    )
    with pd.ExcelWriter(data_dir / "Product.xlsx", engine="openpyxl") as writer:
        product.to_excel(writer, sheet_name="Export_Product", index=False)
    with pd.ExcelWriter(data_dir / "Product_ref.nr.xlsx", engine="openpyxl") as writer:
        refnr.to_excel(writer, sheet_name="Export_Product", index=False)
    workbook_path = tmp_path / "validated.xlsx"
    write_review_workbook(
        workbook_path,
        [
            review_row(
                "REVIEW_001",
                "UNMATCHED_REFNR_001",
                "unmatched_product_refnr",
                "approved_use_corrected_product_ref_nr",
                refnr_identifier="refnr_row_99",
            )
        ],
    )
    application = run_product_refnr_application(
        workbook_path,
        application_output,
        config_dir,
        apply=True,
    )

    result = run_product_materialization(data_dir, workbook_path, application.state_path, output_dir)

    assert result.status == "blocked"
    with result.blockers_path.open(newline="", encoding="utf-8") as handle:
        blocker_types = {row["blocker_type"] for row in csv.DictReader(handle)}
    assert "source_identifier_out_of_range" in blocker_types
