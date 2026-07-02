from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook

from data_ops_lab.product_refnr_decision_validation import run_validate_product_refnr_decisions


HEADERS = [
    "issue_id",
    "issue_type",
    "product_original_identifier",
    "product_refnr_identifier",
    "original_part_nr_sku",
    "corrected_product_ref_nr",
    "optional_pd_ref_nr",
    "match_reason",
    "conflict_reason",
    "risk_explanation",
    "recommended_action",
    "human_decision",
    "human_notes",
]


def add_review_sheet(workbook: Workbook, title: str, rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)


def build_shortlist(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("README").append(["readme"])
    add_review_sheet(
        workbook,
        "Conflicts",
        [
            [
                "CONFLICT_001",
                "conflict",
                "original_row_1",
                "",
                "SKU1",
                "PD2600001",
                "PD2600001",
                "part_nr_sku",
                "multiple",
                "risk",
                "needs_business_context",
                "approved_use_corrected_product_ref_nr",
                "",
            ],
            [
                "CONFLICT_002",
                "conflict",
                "original_row_2",
                "",
                "SKU2",
                "",
                "",
                "part_nr_sku",
                "multiple",
                "risk",
                "needs_business_context",
                "not_allowed",
                "bad value",
            ],
        ],
    )
    add_review_sheet(
        workbook,
        "Unmatched Original Product",
        [
            [
                "UNMATCHED_ORIGINAL_001",
                "unmatched_original_product",
                "original_row_3",
                "",
                "SKU3",
                "",
                "",
                "",
                "no match",
                "risk",
                "approved_create_technical_product_id_only",
                "pending",
                "",
            ]
        ],
    )
    add_review_sheet(
        workbook,
        "Unmatched Product RefNr",
        [
            [
                "UNMATCHED_REFNR_001",
                "unmatched_product_refnr",
                "",
                "refnr_row_4",
                "SKU4",
                "PD2600004",
                "PD2600004",
                "",
                "no original",
                "risk",
                "needs_business_context",
                "approved_use_corrected_product_ref_nr",
                "Include as missing Product row.",
            ]
        ],
    )
    add_review_sheet(
        workbook,
        "Duplicate RefNr Review",
        [
            [
                "DUPLICATE_REFNR_001",
                "duplicate_refnr_review",
                "",
                "refnr_row_5",
                "SKU5",
                "PD2600005",
                "PD2600005",
                "",
                "duplicate",
                "risk",
                "needs_business_context",
                "merge_duplicate_records",
                "Merge into one canonical product.",
            ]
        ],
    )
    workbook.save(path)


def test_validate_product_refnr_decisions_outputs_and_preserves_files(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    shortlist = output_dir / "product_refnr_human_review_shortlist.xlsx"
    build_shortlist(shortlist)

    product = data_dir / "Product.xlsx"
    product_refnr = data_dir / "Product_ref.nr.xlsx"
    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    product.write_bytes(b"product")
    product_refnr.write_bytes(b"refnr")
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    product_before = product.read_bytes()
    refnr_before = product_refnr.read_bytes()
    keys_before = approved_keys.read_text(encoding="utf-8")
    relationships_before = approved_relationships.read_text(encoding="utf-8")

    result = run_validate_product_refnr_decisions(output_dir, shortlist)

    assert result.report_path.exists()
    assert result.summary_csv_path.exists()
    assert result.total_decisions == 5
    assert result.valid_decisions == 4
    assert result.pending_decisions == 1
    assert result.invalid_decisions == 1
    assert result.missing_notes == 1
    assert "Do not apply final decisions yet" in result.recommended_next_step
    assert product.read_bytes() == product_before
    assert product_refnr.read_bytes() == refnr_before
    assert approved_keys.read_text(encoding="utf-8") == keys_before
    assert approved_relationships.read_text(encoding="utf-8") == relationships_before

    report = result.report_path.read_text(encoding="utf-8")
    assert "CONFLICT_002" in report
    assert "CONFLICT_001" in report
    assert "UNMATCHED_ORIGINAL_001" in report

    with result.summary_csv_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert any(row["issue_type"] == "conflict" and row["invalid_decision_count"] == "1" for row in rows)
    assert any(row["issue_type"] == "conflict" and row["missing_notes_count"] == "1" for row in rows)
