from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

from data_ops_lab.product_refnr_final_review_spreadsheet import run_product_refnr_final_review_spreadsheet


HEADERS = [
    "issue_id",
    "issue_type",
    "product_original_identifier",
    "product_refnr_identifier",
    "original_part_nr_sku",
    "corrected_product_ref_nr",
    "optional_pd_ref_nr",
    "match_reason",
    "conflict_reason",
    "risk_explanation",
    "recommended_action",
    "human_decision",
    "human_notes",
]


def add_review_sheet(workbook: Workbook, title: str, rows: list[list[str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(HEADERS)
    for row in rows:
        sheet.append(row)


def build_shortlist(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("README").append(["readme"])
    add_review_sheet(
        workbook,
        "Conflicts",
        [
            [
                "CONFLICT_001",
                "conflict",
                "original_row_1",
                "",
                "SKU1",
                "PD2600001",
                "PD2600001",
                "part_nr_sku",
                "multiple",
                "risk",
                "needs_business_context",
                "approved_use_corrected_product_ref_nr",
                "",
            ],
            [
                "CONFLICT_002",
                "conflict",
                "original_row_2",
                "",
                "SKU2",
                "PD2600002",
                "PD2600002",
                "part_nr_sku",
                "multiple",
                "risk",
                "needs_business_context",
                "rejected",
                "Not valid.",
            ],
        ],
    )
    add_review_sheet(
        workbook,
        "Unmatched Original Product",
        [
            [
                "UNMATCHED_ORIGINAL_001",
                "unmatched_original_product",
                "original_row_3",
                "",
                "SKU3",
                "",
                "",
                "",
                "no match",
                "risk",
                "approved_create_technical_product_id_only",
                "approved_create_technical_product_id_only",
                "",
            ]
        ],
    )
    add_review_sheet(
        workbook,
        "Unmatched Product RefNr",
        [
            [
                "UNMATCHED_REFNR_001",
                "unmatched_product_refnr",
                "",
                "refnr_row_4",
                "SKU4",
                "PD2600004",
                "PD2600004",
                "",
                "no original",
                "risk",
                "needs_business_context",
                "approved_use_corrected_product_ref_nr",
                "Include this Product_ref.nr row.",
            ]
        ],
    )
    add_review_sheet(
        workbook,
        "Duplicate RefNr Review",
        [
            [
                "DUPLICATE_REFNR_001",
                "duplicate_refnr_review",
                "",
                "refnr_row_5",
                "SKU5",
                "PD2600005",
                "PD2600005",
                "",
                "duplicate",
                "risk",
                "needs_business_context",
                "approved_use_corrected_product_ref_nr",
                "",
            ]
        ],
    )
    workbook.save(path)


def test_product_refnr_final_review_spreadsheet_outputs_and_preserves_files(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    shortlist = output_dir / "product_refnr_human_review_shortlist.xlsx"
    build_shortlist(shortlist)

    product = data_dir / "Product.xlsx"
    product_refnr = data_dir / "Product_ref.nr.xlsx"
    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    product.write_bytes(b"product")
    product_refnr.write_bytes(b"refnr")
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    product_before = product.read_bytes()
    refnr_before = product_refnr.read_bytes()
    keys_before = approved_keys.read_text(encoding="utf-8")
    relationships_before = approved_relationships.read_text(encoding="utf-8")

    result = run_product_refnr_final_review_spreadsheet(output_dir, shortlist)

    assert result.xlsx_path.exists()
    assert result.csv_path.exists()
    assert result.readme_path.exists()
    assert result.required_review_count == 2
    assert result.missing_notes_count == 2
    assert result.inconsistency_count == 1
    assert result.all_exceptions_count == 5
    assert product.read_bytes() == product_before
    assert product_refnr.read_bytes() == refnr_before
    assert approved_keys.read_text(encoding="utf-8") == keys_before
    assert approved_relationships.read_text(encoding="utf-8") == relationships_before

    workbook = load_workbook(result.xlsx_path)
    expected = {
        "README",
        "Required Review",
        "Missing Notes",
        "Inconsistencies",
        "All Product Exceptions",
        "Decision Options",
    }
    assert expected.issubset(set(workbook.sheetnames))
    required_headers = [cell.value for cell in workbook["Required Review"][1]]
    assert "final_human_decision" in required_headers
    assert "final_human_notes" in required_headers
    assert workbook["Required Review"].max_row == 3
    assert workbook["All Product Exceptions"].max_row == 6

    with result.csv_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 2
    assert all(row["problem_type"] != "ready" for row in rows)

    readme = result.readme_path.read_text(encoding="utf-8")
    assert "Required Review rows: 2" in readme
