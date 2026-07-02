from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from data_ops_lab.product_refnr_human_review import run_product_refnr_human_review


def build_reconciliation_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    conflicts = workbook.create_sheet("conflicts")
    conflicts.append(
        [
            "original_source_row_number",
            "original_part_nr_sku",
            "original_description",
            "match_method",
            "candidate_count",
            "conflict_reason",
        ]
    )
    conflicts.append([10, "SKU-CONFLICT", "Conflict product", "part_nr_sku", 2, "multiple_product_refnr_rows_matched_original_product"])

    unmatched_original = workbook.create_sheet("unmatched_original_product")
    unmatched_original.append(["original_source_row_number", "original_part_nr_sku", "original_description"])
    unmatched_original.append([20, "SKU-ORIGINAL-ONLY", "Original only"])

    unmatched_refnr = workbook.create_sheet("unmatched_product_refnr")
    unmatched_refnr.append(["refnr_source_row_number", "refnr_product_ref_nr", "refnr_part_nr_sku", "refnr_description"])
    unmatched_refnr.append([30, "PD2600001", "SKU-REFNR-ONLY", "Refnr only"])

    duplicates = workbook.create_sheet("duplicates_in_product_refnr")
    duplicates.append(
        [
            "duplicate_column",
            "duplicate_type",
            "duplicate_value",
            "duplicate_count",
            "refnr_source_row_number",
            "refnr_product_ref_nr",
            "refnr_part_nr_sku",
            "refnr_description",
        ]
    )
    duplicates.append(["part_nr_sku", "part_nr_sku", "SKU-DUP", 2, 40, "PD2600002", "SKU-DUP", "Duplicate row"])

    workbook.save(path)


def test_product_refnr_human_review_shortlist_outputs_and_preserves_files(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    workbook_path = output_dir / "product_refnr_reconciliation_review.xlsx"
    build_reconciliation_workbook(workbook_path)

    product = data_dir / "Product.xlsx"
    product_refnr = data_dir / "Product_ref.nr.xlsx"
    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    product.write_bytes(b"raw product")
    product_refnr.write_bytes(b"raw refnr")
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    product_before = product.read_bytes()
    refnr_before = product_refnr.read_bytes()
    keys_before = approved_keys.read_text(encoding="utf-8")
    relationships_before = approved_relationships.read_text(encoding="utf-8")

    result = run_product_refnr_human_review(output_dir, workbook_path)

    assert result.shortlist_xlsx.exists()
    assert result.shortlist_md.exists()
    assert result.modeling_recommendation_path.exists()
    assert result.conflict_count == 1
    assert result.unmatched_original_count == 1
    assert result.unmatched_refnr_count == 1
    assert result.duplicate_count == 1
    assert result.product_finalized is False
    assert product.read_bytes() == product_before
    assert product_refnr.read_bytes() == refnr_before
    assert approved_keys.read_text(encoding="utf-8") == keys_before
    assert approved_relationships.read_text(encoding="utf-8") == relationships_before

    shortlist = load_workbook(result.shortlist_xlsx)
    expected_sheets = {
        "README",
        "Conflicts",
        "Unmatched Original Product",
        "Unmatched Product RefNr",
        "Duplicate RefNr Review",
        "Decision Options",
    }
    assert expected_sheets.issubset(set(shortlist.sheetnames))
    conflict_headers = [cell.value for cell in shortlist["Conflicts"][1]]
    assert "issue_id" in conflict_headers
    assert "human_decision" in conflict_headers
    assert shortlist["Conflicts"]["L2"].value == "pending"
    options = [cell.value for cell in shortlist["Decision Options"]["A"][1:]]
    assert "approved_use_corrected_product_ref_nr" in options
    assert "pending" in options

    report = result.shortlist_md.read_text(encoding="utf-8")
    assert "Conflicts: 1" in report
    assert "Unmatched original Product rows: 1" in report
    assert "Duplicate Product_ref.nr review rows: 1" in report

    modeling = result.modeling_recommendation_path.read_text(encoding="utf-8")
    assert "`product_id` as technical/generated primary key" in modeling
    assert "Final approval remains pending" in modeling
