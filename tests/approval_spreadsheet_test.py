from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from data_ops_lab.approval_spreadsheet import run_approval_spreadsheet


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def seed_step3d(step3d_dir: Path) -> None:
    write_csv(
        step3d_dir / "serial_aware_key_review.csv",
        [
            {
                "table_name": "salesorder",
                "table_type": "header",
                "source_file": "SalesOrder.csv",
                "candidate_key": "ref_nr",
                "key_type": "natural",
                "non_null_rate": "100.0",
                "uniqueness_rate": "100.0",
                "duplicate_count": "0",
                "expected_prefix": "OC",
                "prefix_match_rate": "100.0",
                "regex_match_rate": "100.0",
                "semantic_namespace": "sales_order",
                "semantic_ref_name": "oc_ref_nr",
                "serial_validation_status": "pending_review",
                "technical_confidence": "high",
                "serial_confidence": "high",
                "combined_confidence": "high",
                "recommended_human_decision": "approve_as_semantic_primary_key",
                "status": "pending_review",
                "notes": "No approval applied.",
            },
            {
                "table_name": "salesorderline",
                "table_type": "line",
                "source_file": "SalesOrderLine.csv",
                "candidate_key": "ref_nr + row_position",
                "key_type": "composite",
                "non_null_rate": "100.0",
                "uniqueness_rate": "100.0",
                "duplicate_count": "0",
                "expected_prefix": "OC",
                "prefix_match_rate": "100.0",
                "regex_match_rate": "100.0",
                "semantic_namespace": "sales_order",
                "semantic_ref_name": "oc_ref_nr",
                "serial_validation_status": "pending_review",
                "technical_confidence": "high",
                "serial_confidence": "high",
                "combined_confidence": "high",
                "recommended_human_decision": "approve_as_technical_key_only",
                "status": "pending_review",
                "notes": "No approval applied.",
            },
            {
                "table_name": "product_export_product",
                "table_type": "master",
                "source_file": "Product.xlsx",
                "candidate_key": "part_nr_sku",
                "key_type": "natural",
                "non_null_rate": "100.0",
                "uniqueness_rate": "99.0",
                "duplicate_count": "1",
                "expected_prefix": "PD",
                "prefix_match_rate": "0.0",
                "regex_match_rate": "0.0",
                "semantic_namespace": "product",
                "semantic_ref_name": "pd_ref_nr",
                "serial_validation_status": "needs_business_context",
                "technical_confidence": "needs_business_context",
                "serial_confidence": "needs_business_context",
                "combined_confidence": "needs_business_context",
                "recommended_human_decision": "needs_business_context",
                "status": "needs_business_context",
                "notes": "No approval applied.",
            },
        ],
    )
    write_csv(
        step3d_dir / "serial_aware_relationship_review.csv",
        [
            {
                "source_table": "salesorderline",
                "source_column": "ref_nr",
                "target_table": "salesorder",
                "target_column": "ref_nr",
                "relationship_type": "header_line",
                "source_semantic_namespace": "sales_order",
                "target_semantic_namespace": "sales_order",
                "expected_prefix": "OC",
                "match_rate": "100.0",
                "unmatched_count": "0",
                "target_duplicate_count": "0",
                "join_risk": "low",
                "prefix_consistency": "consistent",
                "combined_confidence": "high",
                "recommended_human_decision": "approve_header_line_relationship",
                "status": "pending_review",
                "notes": "No approval applied.",
            }
        ],
    )
    (step3d_dir / "human_decision_shortlist.md").write_text("# Shortlist\n", encoding="utf-8")
    (step3d_dir / "conflict_investigation.md").write_text(
        "# Step 3D Conflict Investigation\n\n"
        "## salesorderline2_export_salesorderline\n\n"
        "- Source file: SalesOrderLine2.xlsx\n"
        "- Columns available: ref_nr, creditor\n"
        "- Row count: 10\n"
        "- Ref nr detected prefixes: CR:10\n"
        "- Expected prefix: OC\n"
        "- Detected prefix: CR:10\n"
        "- Technical hypothesis: observed prefix does not match the current semantic mapping for this table name.\n"
        "- Risk: high for automatic relationship/key approval.\n"
        "- Question for human review: should this table be reinterpreted as another document/domain?\n"
        "- recommended_human_decision: needs_business_context\n",
        encoding="utf-8",
    )


def test_approval_spreadsheet_outputs_and_sheets(tmp_path: Path) -> None:
    step3d_dir = tmp_path / "step3d"
    output_dir = tmp_path / "step3e"
    config_dir = tmp_path / "config" / "data_model"
    config_dir.mkdir(parents=True)
    seed_step3d(step3d_dir)

    result = run_approval_spreadsheet(step3d_dir, output_dir, config_dir)

    assert result.decision_count == 5
    assert result.primary_key_count == 1
    assert result.relationship_count == 1
    assert result.technical_key_count == 1
    assert result.needs_context_count == 1
    assert result.conflict_count == 1
    assert result.xlsx_path.exists()
    assert result.csv_path.exists()
    assert (output_dir / "human_approval_matrix_readme.md").exists()

    workbook = load_workbook(result.xlsx_path)
    expected_sheets = {
        "README",
        "Approval Matrix",
        "Primary Keys",
        "Relationships",
        "Technical Line Keys",
        "Needs Context",
        "Conflicts",
        "Decision Options",
    }
    assert expected_sheets.issubset(set(workbook.sheetnames))
    headers = [cell.value for cell in workbook["Approval Matrix"][1]]
    assert "human_decision" in headers
    options = [cell.value for cell in workbook["Decision Options"]["A"][1:]]
    assert "approved" in options
    assert "pending" in options


def test_approval_spreadsheet_preserves_approved_files_and_originaldatabase(tmp_path: Path) -> None:
    step3d_dir = tmp_path / "step3d"
    output_dir = tmp_path / "step3e"
    config_dir = tmp_path / "config" / "data_model"
    original_dir = tmp_path / "originaldatabase"
    config_dir.mkdir(parents=True)
    original_dir.mkdir()
    seed_step3d(step3d_dir)
    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    source_file = original_dir / "raw.csv"
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    source_file.write_text("a,b\n1,2\n", encoding="utf-8")
    keys_before = approved_keys.read_text(encoding="utf-8")
    relationships_before = approved_relationships.read_text(encoding="utf-8")
    source_before = source_file.read_text(encoding="utf-8")

    run_approval_spreadsheet(step3d_dir, output_dir, config_dir)

    assert approved_keys.read_text(encoding="utf-8") == keys_before
    assert approved_relationships.read_text(encoding="utf-8") == relationships_before
    assert source_file.read_text(encoding="utf-8") == source_before
