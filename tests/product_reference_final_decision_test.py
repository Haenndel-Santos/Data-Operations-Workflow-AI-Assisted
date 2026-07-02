from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from data_ops_lab.product_reference_final_decision import run_product_reference_final_decision


def build_review_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)

    duplicate = workbook.create_sheet("duplicate_part_nr_sku")
    duplicate.append(
        [
            "duplicate_group_id",
            "source_row_number",
            "part_nr_sku",
            "classification_from_audit",
            "human_decision",
            "human_notes",
        ]
    )
    duplicate.append(
        [
            "product_dup_001_hash",
            2,
            "RAW-DUP-001",
            "likely_distinct_products_sharing_reference",
            "distinct_products_same_reference",
            "Confirmed as two distinct products.",
        ]
    )
    duplicate.append(
        [
            "product_dup_002_hash",
            3,
            "RAW-DUP-002",
            "likely_same_product_duplicate_record",
            "same_product_duplicate_record",
            "Same product listed twice.",
        ]
    )

    empty = workbook.create_sheet("empty_part_nr_sku")
    empty.append(["source_row_number", "current_audit_classification", "human_decision", "human_notes"])
    empty.append([10, "repair_candidate", "repair_reference", "Find missing SKU from supplier file."])
    empty.append([11, "requires_human_review", "", "Waiting for product owner."])

    non_pd = workbook.create_sheet("non_pd_pattern_products")
    non_pd.append(["source_row_number", "part_nr_sku", "human_decision", "human_notes"])
    non_pd.append([20, "TEXTUAL-REF", "valid_textual_product_reference", "Valid legacy reference."])
    non_pd.append([21, "TEXTUAL-REVIEW", "requires_more_investigation", "Unclear status."])

    log = workbook.create_sheet("decision_log")
    log.append(["decision_date", "reviewer", "topic", "decision", "rationale", "follow_up_action"])
    log.append(["2026-06-17", "Business User", "Product key", "Use generated product_id", "Duplicate refs exist", "Clean duplicates"])

    workbook.save(path)


def test_product_reference_final_decision_reads_human_decisions_and_preserves_approved_files(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    config_dir.mkdir(parents=True)
    workbook_path = output_dir / "product_reference_human_review.xlsx"
    build_review_workbook(workbook_path)

    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    workbook_before = workbook_path.read_bytes()
    keys_before = approved_keys.read_text(encoding="utf-8")
    relationships_before = approved_relationships.read_text(encoding="utf-8")

    result = run_product_reference_final_decision(output_dir, workbook_path)

    assert result.report_path.exists()
    assert result.total_reviewed_rows == 6
    assert result.unresolved_rows == 2
    assert result.more_investigation_rows == 1
    assert result.distinct_same_reference_rows == 1
    assert result.part_nr_sku_unique_key_recommended is False
    assert workbook_path.read_bytes() == workbook_before
    assert approved_keys.read_text(encoding="utf-8") == keys_before
    assert approved_relationships.read_text(encoding="utf-8") == relationships_before

    report = result.report_path.read_text(encoding="utf-8")
    assert "distinct_products_same_reference" in report
    assert "valid_textual_product_reference" in report
    assert "repair_reference" in report
    assert "Can `part_nr_sku` be used as the Product primary key? No" in report
    assert "`part_nr_sku` must not be approved as the Product primary key; use generated `product_id` instead." in report
    assert "`product_id`: generated technical primary key" in report
    assert "`pd_ref_nr` should remain optional" in report

    reopened = load_workbook(workbook_path)
    assert "duplicate_part_nr_sku" in reopened.sheetnames
