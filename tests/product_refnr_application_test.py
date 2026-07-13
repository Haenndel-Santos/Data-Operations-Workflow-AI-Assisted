from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from data_ops_lab.product_refnr_application import run_product_refnr_application
from data_ops_lab.product_refnr_final_review_spreadsheet import REVIEW_COLUMNS


def review_row(review_id: str, issue_type: str, decision: str, notes: str) -> dict[str, str]:
    return {
        "review_id": review_id,
        "original_sheet": "Conflicts",
        "original_excel_row": "2",
        "issue_id": review_id.replace("REVIEW", "ISSUE"),
        "issue_type": issue_type,
        "product_original_identifier": "private-original-value",
        "product_refnr_identifier": "private-refnr-value",
        "original_part_nr_sku": "private-sku-value",
        "corrected_product_ref_nr": "private-corrected-value",
        "optional_pd_ref_nr": "private-pd-value",
        "current_human_decision": decision,
        "current_human_notes": notes,
        "problem_type": "",
        "problem_explanation": "",
        "suggested_human_decision": decision,
        "suggested_human_notes": notes,
        "required_action": "",
        "final_human_decision": decision,
        "final_human_notes": notes,
    }


def build_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("README").append(["readme"])
    for title in ("Required Review", "Missing Notes", "Inconsistencies", "All Product Exceptions"):
        sheet = workbook.create_sheet(title)
        sheet.append(REVIEW_COLUMNS)
        if title in {"Required Review", "All Product Exceptions"}:
            for row in rows:
                sheet.append([row.get(column, "") for column in REVIEW_COLUMNS])
    workbook.save(path)


def test_product_application_dry_run_preserves_config_and_sources(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    config_dir = tmp_path / "config" / "data_model"
    data_dir = tmp_path / "originaldatabase"
    output_dir.mkdir()
    config_dir.mkdir(parents=True)
    data_dir.mkdir()
    workbook_path = tmp_path / "validated.xlsx"
    build_workbook(
        workbook_path,
        [
            review_row("REVIEW_001", "conflict", "approved_use_corrected_product_ref_nr", "Approved."),
            review_row("REVIEW_002", "unmatched_original_product", "rejected", "Invalid Product record."),
        ],
    )

    product = data_dir / "Product.xlsx"
    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    product.write_bytes(b"protected-product")
    approved_keys.write_text("approved_keys: []\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships: []\n", encoding="utf-8")
    protected = {path: path.read_bytes() for path in (product, approved_keys, approved_relationships, workbook_path)}

    result = run_product_refnr_application(workbook_path, output_dir, config_dir)

    assert result.dry_run is True
    assert result.state_changed is False
    assert result.total_decisions == 2
    assert result.approved_decisions == 1
    assert result.rejected_decisions == 1
    assert result.report_path.exists()
    assert result.plan_csv_path.exists()
    assert not result.state_path.exists()
    assert all(path.read_bytes() == content for path, content in protected.items())
    assert "Mode: `dry-run`" in result.report_path.read_text(encoding="utf-8")


def test_product_application_apply_is_private_and_idempotent(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    config_dir.mkdir(parents=True)
    workbook_path = tmp_path / "validated.xlsx"
    build_workbook(
        workbook_path,
        [
            review_row("REVIEW_001", "conflict", "approved_use_corrected_product_ref_nr", "Approved."),
            review_row("REVIEW_002", "unmatched_original_product", "rejected", "Invalid Product record."),
        ],
    )

    first = run_product_refnr_application(workbook_path, output_dir, config_dir, apply=True)
    state_before = first.state_path.read_bytes()
    second = run_product_refnr_application(workbook_path, output_dir, config_dir, apply=True)
    state = yaml.safe_load(first.state_path.read_text(encoding="utf-8"))

    assert first.state_changed is True
    assert second.state_changed is False
    assert second.decision_digest == first.decision_digest
    assert first.state_path.read_bytes() == state_before
    assert state["model_contract"]["primary_key"] == "product_id"
    assert state["model_contract"]["business_reference"] == "part_nr_sku"
    assert state["counts"] == {"total": 2, "approved": 1, "rejected": 1}
    rejected = next(item for item in state["decisions"] if item["decision"] == "rejected")
    assert rejected["action"] == "exclude_from_target_product_model"
    state_text = first.state_path.read_text(encoding="utf-8")
    assert "private-sku-value" not in state_text
    assert "private-corrected-value" not in state_text


def test_product_application_refuses_unresolved_context(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    workbook_path = tmp_path / "validated.xlsx"
    build_workbook(
        workbook_path,
        [review_row("REVIEW_001", "conflict", "needs_business_context", "Owner review still required.")],
    )

    with pytest.raises(ValueError, match="still needs business context"):
        run_product_refnr_application(workbook_path, output_dir, config_dir, apply=True)

    assert not (config_dir / "product_reconciliation_state.yml").exists()


def test_product_application_refuses_different_existing_state(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    config_dir.mkdir(parents=True)
    workbook_path = tmp_path / "validated.xlsx"
    build_workbook(
        workbook_path,
        [review_row("REVIEW_001", "conflict", "approved_use_corrected_product_ref_nr", "Approved.")],
    )
    run_product_refnr_application(workbook_path, output_dir, config_dir, apply=True)
    state_path = config_dir / "product_reconciliation_state.yml"
    state_before = state_path.read_bytes()

    replacement_workbook = tmp_path / "replacement.xlsx"
    build_workbook(
        replacement_workbook,
        [review_row("REVIEW_001", "conflict", "rejected", "Invalid Product record.")],
    )

    with pytest.raises(ValueError, match="different Product reconciliation state"):
        run_product_refnr_application(replacement_workbook, output_dir, config_dir, apply=True)

    assert state_path.read_bytes() == state_before


def test_product_application_refuses_existing_state_without_digest(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    config_dir.mkdir(parents=True)
    state_path = config_dir / "product_reconciliation_state.yml"
    state_path.write_text("version: 1\nstatus: unknown\n", encoding="utf-8")
    workbook_path = tmp_path / "validated.xlsx"
    build_workbook(
        workbook_path,
        [review_row("REVIEW_001", "conflict", "approved_use_corrected_product_ref_nr", "Approved.")],
    )

    with pytest.raises(ValueError, match="different Product reconciliation state"):
        run_product_refnr_application(workbook_path, output_dir, config_dir, apply=True)

    assert state_path.read_text(encoding="utf-8") == "version: 1\nstatus: unknown\n"
