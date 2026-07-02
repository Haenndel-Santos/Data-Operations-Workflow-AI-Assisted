from __future__ import annotations

from pathlib import Path

import pandas as pd
import yaml
from openpyxl import Workbook, load_workbook

from data_ops_lab.canonical_model import run_canonical_model_alignment


def test_canonical_model_alignment_product_rules_and_preserved_files(tmp_path: Path) -> None:
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    output_dir = tmp_path / "outputs"
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    output_dir.mkdir(parents=True)

    product_path = data_dir / "Product.xlsx"
    product_rows = pd.DataFrame(
        [
            {
                "Part. nr. (SKU)": "PD2600001",
                "Name": "Serial Product",
                "Item Description": "Same description",
            },
            {
                "Part. nr. (SKU)": "LOCAL-REF-1",
                "Name": "Functional Product",
                "Item Description": "Local only",
            },
            {
                "Part. nr. (SKU)": "DUP-REF",
                "Name": "Duplicate A",
                "Item Description": "First record",
            },
            {
                "Part. nr. (SKU)": "DUP-REF",
                "Name": "Duplicate B",
                "Item Description": "Different record",
            },
            {
                "Part. nr. (SKU)": "",
                "Name": "No Reference",
                "Item Description": "Blank reference",
            },
        ]
    )
    with pd.ExcelWriter(product_path, engine="openpyxl") as writer:
        product_rows.to_excel(writer, sheet_name="Export_Product", index=False)
    workbook = Workbook()
    workbook.active.title = "Approval Matrix"
    workbook.active.append(["table_name", "source_table", "target_table"])
    workbook.active.append(["product_export_product", "", ""])
    workbook.save(output_dir / "human_approval_matrix.xlsx")
    (output_dir / "human_approval_matrix.csv").write_text(
        "table_name,source_table,target_table\n"
        "product_export_product,,\n"
        "salesorderline_export_salesorderline,,\n",
        encoding="utf-8",
    )

    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    product_before = product_path.read_bytes()
    keys_before = approved_keys.read_text(encoding="utf-8")
    relationships_before = approved_relationships.read_text(encoding="utf-8")

    result = run_canonical_model_alignment(data_dir, config_dir, output_dir)

    assert result.canonical_count == 14
    assert result.complement_count == 1
    assert result.canonical_review_xlsx.exists()
    assert result.product_status == "manually_confirmed_pending_duplicate_validation"
    assert result.organisation_status == "needs_business_context"
    assert product_path.read_bytes() == product_before
    assert approved_keys.read_text(encoding="utf-8") == keys_before
    assert approved_relationships.read_text(encoding="utf-8") == relationships_before

    canonical_tables = yaml.safe_load((config_dir / "canonical_tables.yml").read_text(encoding="utf-8"))
    product = canonical_tables["canonical_tables"]["product"]
    assert product["semantic_ref_name"] == "product_ref"
    assert product["optional_serial_ref_name"] == "pd_ref_nr"
    assert product["primary_key_rule"] == "part_nr_sku_business_reference"
    assert product["optional_serial_rule"] == "PDYY99999 when available"
    assert product["status"] == "manually_confirmed_pending_duplicate_validation"
    assert canonical_tables["canonical_tables"]["organisation"]["expected_prefix"] is None
    assert canonical_tables["canonical_tables"]["organisation"]["status"] == "needs_business_context"

    business_rules = yaml.safe_load((config_dir / "business_rules.yml").read_text(encoding="utf-8"))
    product_rules = "\n".join(business_rules["product_rules"])
    assert "Product canonical reference is part_nr_sku." in product_rules
    assert "PD serial reference is optional, not mandatory." in product_rules
    assert "Products without PD code are valid product records." in product_rules
    assert "pd_ref_nr should only be derived when the value matches PD serial format." in product_rules

    report = (output_dir / "product_reference_validation.md").read_text(encoding="utf-8")
    assert "Total products: 5" in report
    assert "`part_nr_sku` matching `PDYY99999`: 1" in report
    assert "`part_nr_sku` not matching PD pattern: 3" in report
    assert "Duplicate count for `part_nr_sku`: 1" in report
    assert "product_dup_001_" in report
    assert "DUP-REF" not in report
    assert "LOCAL-REF-1" not in report

    canonical_review = load_workbook(result.canonical_review_xlsx)
    assert "Canonical Model" in canonical_review.sheetnames
    assert "Complement Tables" in canonical_review.sheetnames
    assert "Product Validation" in canonical_review.sheetnames
    complement_rows = list(canonical_review["Complement Tables"].iter_rows(values_only=True))
    assert ("salesorderline_export_salesorderline", "document_line", "pending_human_review") in complement_rows
