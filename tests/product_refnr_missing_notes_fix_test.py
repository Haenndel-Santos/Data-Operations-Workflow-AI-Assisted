from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook

from data_ops_lab.product_refnr_final_review_spreadsheet import REVIEW_COLUMNS
from data_ops_lab.product_refnr_missing_notes_fix import run_product_refnr_missing_notes_fix


def final_review_row(
    review_id: str,
    issue_type: str,
    decision: str,
    notes: str,
    original_sheet: str = "Required Review",
) -> dict[str, str]:
    return {
        "review_id": review_id,
        "original_sheet": original_sheet,
        "original_excel_row": "2",
        "issue_id": review_id.replace("REVIEW", "ISSUE"),
        "issue_type": issue_type,
        "product_original_identifier": "original_row_1",
        "product_refnr_identifier": "refnr_row_1",
        "original_part_nr_sku": "SKU1",
        "corrected_product_ref_nr": "PD2600001",
        "optional_pd_ref_nr": "PD2600001",
        "current_human_decision": decision,
        "current_human_notes": notes,
        "problem_type": "missing_note" if not notes else "ready",
        "problem_explanation": "Requires final note." if not notes else "No blocker.",
        "suggested_human_decision": decision,
        "suggested_human_notes": "",
        "required_action": "Fill final_human_notes." if not notes else "No action required.",
        "final_human_decision": decision,
        "final_human_notes": notes,
    }


def add_review_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(REVIEW_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in REVIEW_COLUMNS])


def build_final_review(path: Path) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.create_sheet("README").append(["readme"])
    missing_conflict = final_review_row("REVIEW_001", "conflict", "approved_use_corrected_product_ref_nr", "")
    missing_unmatched = final_review_row(
        "REVIEW_002",
        "unmatched_original_product",
        "approved_create_technical_product_id_only",
        "",
    )
    complete = final_review_row(
        "REVIEW_003",
        "unmatched_product_refnr",
        "approved_use_corrected_product_ref_nr",
        "Already has a note.",
    )
    add_review_sheet(workbook, "Required Review", [missing_conflict, missing_unmatched, complete])
    add_review_sheet(workbook, "Missing Notes", [missing_conflict])
    add_review_sheet(workbook, "Inconsistencies", [])
    add_review_sheet(workbook, "All Product Exceptions", [missing_conflict, missing_unmatched, complete])
    workbook.save(path)


def test_product_refnr_missing_notes_fix_outputs_only_missing_notes_and_preserves_files(tmp_path: Path) -> None:
    output_dir = tmp_path / "outputs"
    data_dir = tmp_path / "originaldatabase"
    config_dir = tmp_path / "config" / "data_model"
    output_dir.mkdir()
    data_dir.mkdir()
    config_dir.mkdir(parents=True)
    workbook_path = output_dir / "product_refnr_final_review_required.xlsx"
    report_path = output_dir / "product_refnr_final_review_validation_report.md"
    summary_path = output_dir / "product_refnr_final_review_validation_summary.csv"
    build_final_review(workbook_path)
    report_path.write_text("# report\n", encoding="utf-8")
    summary_path.write_text("issue_type,total_rows,missing_notes_count\nconflict,1,1\n", encoding="utf-8")

    product = data_dir / "Product.xlsx"
    product_refnr = data_dir / "Product_ref.nr.xlsx"
    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    product.write_bytes(b"product")
    product_refnr.write_bytes(b"refnr")
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    protected_before = {
        product: product.read_bytes(),
        product_refnr: product_refnr.read_bytes(),
        approved_keys: approved_keys.read_text(encoding="utf-8"),
        approved_relationships: approved_relationships.read_text(encoding="utf-8"),
    }

    result = run_product_refnr_missing_notes_fix(output_dir, workbook_path, report_path, summary_path)

    assert result.missing_notes_count == 2
    assert result.xlsx_path.exists()
    assert result.csv_path.exists()
    assert result.readme_path.exists()
    assert product.read_bytes() == protected_before[product]
    assert product_refnr.read_bytes() == protected_before[product_refnr]
    assert approved_keys.read_text(encoding="utf-8") == protected_before[approved_keys]
    assert approved_relationships.read_text(encoding="utf-8") == protected_before[approved_relationships]

    workbook = load_workbook(result.xlsx_path)
    assert {"README", "Missing Notes Fix", "Suggested Notes", "Decision Options"}.issubset(set(workbook.sheetnames))
    sheet = workbook["Missing Notes Fix"]
    headers = [cell.value for cell in sheet[1]]
    assert sheet.max_row == 3
    assert "final_human_notes_to_apply" in headers
    suggestion_col = headers.index("suggested_final_human_notes") + 1
    suggestions = [sheet.cell(row=row, column=suggestion_col).value for row in range(2, sheet.max_row + 1)]
    assert "Corrected product reference accepted after manual review." in suggestions
    assert "No reliable natural product reference confirmed; product retained with generated technical product_id only." in suggestions
    notes_col = headers.index("current_final_human_notes") + 1
    assert all(not sheet.cell(row=row, column=notes_col).value for row in range(2, sheet.max_row + 1))

    with result.csv_path.open(newline="", encoding="utf-8") as handle:
        rows = list(csv.DictReader(handle))
    assert len(rows) == 2
    assert all(row["current_final_human_notes"] == "" for row in rows)
    assert {row["fix_id"] for row in rows} == {"MISSING_NOTE_001", "MISSING_NOTE_002"}
