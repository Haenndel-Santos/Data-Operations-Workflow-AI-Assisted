from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from data_ops_lab.product_refnr_reconciliation import run_product_refnr_reconciliation


def test_product_refnr_reconciliation_outputs_and_preserves_approved_files(tmp_path: Path) -> None:
    db_dir = tmp_path / "db"
    data_dir = tmp_path / "originaldatabase"
    output_dir = tmp_path / "outputs"
    config_dir = tmp_path / "config" / "data_model"
    db_dir.mkdir()
    data_dir.mkdir()
    output_dir.mkdir()
    config_dir.mkdir(parents=True)

    product_rows = pd.DataFrame(
        [
            {
                "Part nr. (SKU)": "SKU-001",
                "Item description": "Product One",
                "Product group name": "Group A",
                "Name": "Product 1",
            },
            {
                "Part nr. (SKU)": "SKU-002",
                "Item description": "Product Two",
                "Product group name": "Group B",
                "Name": "Product 2",
            },
            {
                "Part nr. (SKU)": "SKU-UNMATCHED",
                "Item description": "Unmatched Original",
                "Product group name": "Group C",
                "Name": "Product 3",
            },
        ]
    )
    product_path = data_dir / "Product.xlsx"
    with pd.ExcelWriter(product_path, engine="openpyxl") as writer:
        product_rows.to_excel(writer, sheet_name="Export_Product", index=False)

    refnr_rows = pd.DataFrame(
        [
            {
                "Ref. nr.": "PD2600001",
                "Part nr. (SKU)": "SKU-001",
                "Item description": "Product One",
                "Product group name": "Group A",
                "Name": "Product 1",
            },
            {
                "Ref. nr.": "PD2600002",
                "Part nr. (SKU)": "SKU-002",
                "Item description": "Product Two",
                "Product group name": "Group B",
                "Name": "Product 2",
            },
            {
                "Ref. nr.": "PD2600002",
                "Part nr. (SKU)": "SKU-REF-DUP",
                "Item description": "Duplicate Ref",
                "Product group name": "Group B",
                "Name": "Product Duplicate",
            },
            {
                "Ref. nr.": "PD2600004",
                "Part nr. (SKU)": "SKU-ONLY-REFNR",
                "Item description": "Only Refnr",
                "Product group name": "Group D",
                "Name": "Product 4",
            },
        ]
    )
    refnr_path = db_dir / "Product_ref.nr.xlsx"
    with pd.ExcelWriter(refnr_path, engine="openpyxl") as writer:
        refnr_rows.to_excel(writer, sheet_name="Export_Product", index=False)

    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    final_decision = output_dir / "product_reference_final_decision.md"
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    final_decision.write_text("previous final decision should not be overwritten\n", encoding="utf-8")
    product_before = product_path.read_bytes()
    refnr_before = refnr_path.read_bytes()
    keys_before = approved_keys.read_text(encoding="utf-8")
    relationships_before = approved_relationships.read_text(encoding="utf-8")
    final_before = final_decision.read_text(encoding="utf-8")

    result = run_product_refnr_reconciliation(db_dir, data_dir, output_dir)

    assert result.source_path == refnr_path
    assert result.product_refnr_rows == 4
    assert result.original_rows == 3
    assert result.matched_rows == 2
    assert result.corrected_refnr_rows == 2
    assert result.unmatched_original_rows == 1
    assert result.unmatched_refnr_rows == 2
    assert result.product_finalized is False
    assert result.workbook_path.exists()
    assert result.report_path.exists()
    assert result.schema_report_path.exists()
    assert product_path.read_bytes() == product_before
    assert refnr_path.read_bytes() == refnr_before
    assert approved_keys.read_text(encoding="utf-8") == keys_before
    assert approved_relationships.read_text(encoding="utf-8") == relationships_before
    assert final_decision.read_text(encoding="utf-8") == final_before

    workbook = load_workbook(result.workbook_path)
    expected_sheets = {
        "matched_products",
        "refnr_corrections",
        "conflicts",
        "unmatched_original_product",
        "unmatched_product_refnr",
        "duplicates_in_product_refnr",
        "reconciliation_summary",
        "decision_log",
    }
    assert expected_sheets.issubset(set(workbook.sheetnames))

    schema_report = result.schema_report_path.read_text(encoding="utf-8")
    assert "Product_ref.nr.xlsx" in schema_report
    assert "`ref_nr`" in schema_report
    assert "`part_nr_sku`" in schema_report

    report = result.report_path.read_text(encoding="utf-8")
    assert "not_finalized_pending_reconciliation_review" in report
    assert "`product_ref_nr`: corrected canonical product reference from `Product_ref.nr`" in report
    assert "Do not update `approved_keys.yml` or `approved_relationships.yml`" in report
