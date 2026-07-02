from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from data_ops_lab.product_reference_audit import run_product_reference_audit
from data_ops_lab.product_reference_review_spreadsheet import run_product_reference_review_spreadsheet


def test_product_reference_review_spreadsheet_contains_raw_values_and_preserves_approved_files(tmp_path: Path) -> None:
    data_dir = tmp_path / "originaldatabase"
    output_dir = tmp_path / "outputs"
    config_dir = tmp_path / "config" / "data_model"
    data_dir.mkdir()
    output_dir.mkdir()
    config_dir.mkdir(parents=True)

    product_path = data_dir / "Product.xlsx"
    rows = pd.DataFrame(
        [
            {
                "Part. nr. (SKU)": "RAW-DUP-001",
                "Suppl. part nr. (SKU)": "SUP-A",
                "Item Description": "Duplicate product A",
                "Name": "Name A",
                "Creditor": "Creditor A",
                "Product Group Name": "Group A",
                "Gross Sales": "100",
            },
            {
                "Part. nr. (SKU)": "RAW-DUP-001",
                "Suppl. part nr. (SKU)": "SUP-B",
                "Item Description": "Duplicate product B",
                "Name": "Name B",
                "Creditor": "Creditor B",
                "Product Group Name": "Group B",
                "Gross Sales": "200",
            },
            {
                "Part. nr. (SKU)": "PD2600001",
                "Suppl. part nr. (SKU)": "",
                "Item Description": "Formal PD product",
                "Name": "Name C",
                "Creditor": "Creditor C",
                "Product Group Name": "Group C",
                "Gross Sales": "300",
            },
            {
                "Part. nr. (SKU)": "TEXTUAL-REF",
                "Suppl. part nr. (SKU)": "",
                "Item Description": "Textual product",
                "Name": "Name D",
                "Creditor": "Creditor D",
                "Product Group Name": "Group D",
                "Gross Sales": "400",
            },
            {
                "Part. nr. (SKU)": "",
                "Suppl. part nr. (SKU)": "SUP-EMPTY",
                "Item Description": "Missing reference product",
                "Name": "Name E",
                "Creditor": "Creditor E",
                "Product Group Name": "Group E",
                "Gross Sales": "500",
            },
        ]
    )
    with pd.ExcelWriter(product_path, engine="openpyxl") as writer:
        rows.to_excel(writer, sheet_name="Export_Product", index=False)

    approved_keys = config_dir / "approved_keys.yml"
    approved_relationships = config_dir / "approved_relationships.yml"
    approved_keys.write_text("approved_keys:\n- keep: true\n", encoding="utf-8")
    approved_relationships.write_text("approved_relationships:\n- keep: true\n", encoding="utf-8")
    product_before = product_path.read_bytes()
    keys_before = approved_keys.read_text(encoding="utf-8")
    relationships_before = approved_relationships.read_text(encoding="utf-8")

    markdown_result = run_product_reference_audit(data_dir, output_dir)
    result = run_product_reference_review_spreadsheet(data_dir, output_dir)

    assert result.xlsx_path.exists()
    assert result.duplicate_rows == 2
    assert result.empty_rows == 1
    assert result.non_pd_rows == 3
    assert product_path.read_bytes() == product_before
    assert approved_keys.read_text(encoding="utf-8") == keys_before
    assert approved_relationships.read_text(encoding="utf-8") == relationships_before

    workbook = load_workbook(result.xlsx_path)
    expected_sheets = {
        "duplicate_part_nr_sku",
        "empty_part_nr_sku",
        "non_pd_pattern_products",
        "review_summary",
        "decision_log",
    }
    assert expected_sheets.issubset(set(workbook.sheetnames))

    duplicate_values = [
        cell.value
        for row in workbook["duplicate_part_nr_sku"].iter_rows(values_only=False)
        for cell in row
    ]
    assert "RAW-DUP-001" in duplicate_values
    assert "human_decision" in duplicate_values
    assert "human_notes" in duplicate_values

    empty_values = [
        cell.value
        for row in workbook["empty_part_nr_sku"].iter_rows(values_only=False)
        for cell in row
    ]
    assert "repair_reference" in empty_values

    non_pd_values = [
        cell.value
        for row in workbook["non_pd_pattern_products"].iter_rows(values_only=False)
        for cell in row
    ]
    assert "TEXTUAL-REF" in non_pd_values

    public_markdown = markdown_result.report_path.read_text(encoding="utf-8")
    assert "RAW-DUP-001" not in public_markdown
